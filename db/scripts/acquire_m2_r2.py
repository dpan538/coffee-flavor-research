#!/usr/bin/env python3
"""Owner-private M2 R2 source adapters with frozen source hashes and typed masks.

Acquisition never grants rights. Only explicitly admitted source-specific tasks
are written. This script does not fit models or inspect evaluation outputs.
"""

from __future__ import annotations

import argparse
import csv
from collections import Counter, defaultdict
from datetime import datetime, timezone
import hashlib
import io
import itertools
import json
import math
from pathlib import Path
import re
import struct
import zipfile

import openpyxl
from lxml import etree

ROOT = Path(__file__).resolve().parents[2]
PUBLIC = ROOT / "db/data/backend-sequential-model-v2/revisions/r2"
PRIVATE = (
    Path.home()
    / "Library/Application Support/Coffee Flavor Research"
    / "backend-sequential-model-v2/revisions/r2"
)
SOURCES = PRIVATE / "sources"
SEED = "M2_R2_SOURCE_SPLIT_20260905"
DIGESTS = {
    "liberica-data.zip": "4d56b9005fb57690d2ccd80f5f62b411ffa17bfcf2a34d6371525bfd9e0428dc",
    "barahona2020.xml": "9bfc04c994bfab6a9a54109ec1f69d3aef174471cb861d647072976cfaf69aef",
    "barahona2020-supp.zip": "db6b73c4ac0a19ed84263a60504eedd8004f7d94f0c4883c96715375efcf37a8",
    "heo2019.xml": "5dbf822580b45d3990b90c0a7bf9f962e1cb9f4af5f4b099909f9751d345e65f",
    "croijmans-coffee-experts.xlsx": "80377fdaced0ada4511e321ef7c40e2b784c3cedeca062889af0376f278a39d8",
    "croijmans-coffee-novices.xlsx": "7ca63baecb3f77cdca4a88780e94dfb81079d35c8fa2db97abe383aa5a09588a",
    "croijmans-coffee-wineexperts.xlsx": "6efe3ba18f841ecf53a1babf33c0d8316f86089225810747d65b9e60f8fc73e5",
    "croijmans-readme.txt": "d3d0cbb48af148503db7d23c27f6734700a69da26c4a8c37d29909f017143cf4",
    "croijmans-dans-api.json": "40048d51dee282ee30ad9bc028c1b74969cc698b14fd013d97b13dd59e79f826",
    "croijmans-stimuli-sorting.xlsx": "ec99a904ce75f70e7fcaec7897484f4940e315b625b77850a3c60dc596bbe9cc",
    "croijmans-stimulus-comparisons.tab": "a8ba7d75bd9bd3ceee2a760a712257ae120d57d257f60b375e232ac0bd0b4c3d",
}


def digest(data):
    return hashlib.sha256(data).hexdigest()


def checked(name):
    data = (SOURCES / name).read_bytes()
    assert digest(data) == DIGESTS[name], f"Source changed: {name}"
    return data


def save(path, value, freeze=False):
    data = (
        json.dumps(value, ensure_ascii=False, sort_keys=True, indent=2) + "\n"
    ).encode()
    if freeze and path.exists():
        assert path.read_bytes() == data, f"Refusing to alter frozen adapter: {path}"
    else:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(data)
        if PRIVATE in path.parents:
            path.chmod(0o600)
    return {"path": str(path), "sha256": digest(data), "bytes": len(data)}


def assigned(identifier, scope):
    value = digest(f"{SEED}|{scope}|{identifier}".encode())
    return "CONFIRMATION" if int(value[:8], 16) % 5 == 0 else "DEVELOPMENT"


def liberica():
    archive = zipfile.ZipFile(io.BytesIO(checked("liberica-data.zip")))
    filename = "Liberica Coffee Sensory/Liberica Coffee Bags/Dataset.xlsx"
    workbook_bytes = archive.read(filename)
    workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    sheet = workbook["RATA Test"]
    # Coordinates identify the original respondent blocks, not the spreadsheet's
    # formulas, summary means, hedonic sheets, or inconsistent SPSS recodes.
    fields = [
        (3, 0, "Green", "green_aroma", "Green Aroma"),
        (3, 12, "Jackfruit", "jackfruit_aroma", "Jackfruit Aroma"),
        (34, 0, "Smoky Aroma", "smoky_aroma", "Smoky Aroma"),
        (34, 12, "Sweet", "sweet", "Sweet"),
        (65, 0, "Sour", "sour", "Sour"),
        (65, 12, "Bitter", "bitter", "Bitter"),
        (96, 0, "Roasty Flavor", "roasty_flavor", "Roasty Flavor"),
        (96, 12, "Bitter Aftertaste", "bitter_aftertaste", "Bitter Aftertaste"),
        (
            127,
            0,
            "Astringent Aftertaste",
            "astringent_aftertaste",
            "Astringent Aftertaste",
        ),
        (127, 12, "Body", "body", "Body"),
    ]
    conditions = ["LD0", "LD5", "LD10", "MD0", "MD5", "MD10", "DD0", "DD5", "DD10"]
    participants = [f"liberica:participant:{i:02d}" for i in range(1, 26)]
    rows = {}
    for pid in participants:
        for condition in conditions:
            rid = f"{pid}|{condition}"
            rows[rid] = {
                "record_id": rid,
                "source_family": "family.meilina_liberica_bags_2025",
                "participant_id": pid,
                "condition_id": condition,
                "coffee_group_id": "liberica:study_materials_lots_undocumented",
                "participant_split": assigned(pid, "liberica_participant"),
                "condition_split": assigned(condition, "liberica_condition"),
                "role": "AUX_COFFEE_WEAK_LABEL",
                "stratum": "LIBERICA_COFFEE_BAGS_WITH_OR_WITHOUT_SMOKED_COFFEE_LEAVES",
                "source_C0": None,
                "source_C1": None,
                "responses": {},
                "response_masks": {},
                "response_states": {},
                "source_cells": {},
            }
    original_initials = {}
    for header, offset, source_label, key, display_label in fields:
        assert sheet.cell(header, offset + 3).value == source_label
        for j, condition in enumerate(conditions):
            assert (
                sheet.cell(header + 1, offset + 3 + j).value == f"{condition} ({j + 1})"
            )
        for i, pid in enumerate(participants, 1):
            row = header + 1 + i
            assert sheet.cell(row, offset + 1).value == i
            initials = sheet.cell(row, offset + 2).value
            if pid in original_initials:
                assert (
                    initials == original_initials[pid]
                ), "Participant alignment changed"
            original_initials[pid] = initials
            for j, condition in enumerate(conditions):
                cell = sheet.cell(row, offset + 3 + j)
                value = cell.value
                assert value is None or (isinstance(value, int) and value in range(6))
                record = rows[f"{pid}|{condition}"]
                record["responses"][key] = value
                record["response_masks"][key] = value is not None
                record["response_states"][key] = (
                    "OBSERVED" if value is not None else "NOT_MEASURED"
                )
                record["source_cells"][key] = f"RATA Test!{cell.coordinate}"
    rows = list(rows.values())
    assert len(rows) == 225 and len({r["record_id"] for r in rows}) == 225
    assert all(len(r["responses"]) == 10 for r in rows)
    assert sum(sum(r["response_masks"].values()) for r in rows) == 2250
    assert all(r["source_C0"] is None and r["source_C1"] is None for r in rows)
    # Native codes have no published anchors in this repository. Zero is an
    # observed category. No code becomes physical absence or scaled intensity.
    contract = {
        "source_id": "MEILINA_LIBERICA_BAGS_2025",
        "doi": "10.17632/m3n2gc4dv6.1",
        "source_url": "https://data.mendeley.com/datasets/m3n2gc4dv6/1",
        "download_url": "https://data.mendeley.com/public-api/zip/m3n2gc4dv6/download/1",
        "authors": ["Lita Meilina", "Rangganis Ulya Auliya", "Niken Widya Palupi"],
        "license": "CC-BY-4.0",
        "rights_scope": "NONCOMMERCIAL_RESEARCH_USE",
        "attribution": "Meilina, L.; Ulya Auliya, R.; Widya Palupi, N. (2025). Liberica Coffee Sensory. Mendeley Data V1. doi:10.17632/m3n2gc4dv6.1. Modified: respondent sensory-block extraction and pseudonymous IDs.",
        "license_evidence": "Source landing page: Licence CC BY 4.0; original cached liberica-mendeley.html.",
        "raw_sha256": DIGESTS["liberica-data.zip"],
        "workbook_sha256": digest(workbook_bytes),
        "member": filename,
        "admitted_task": "RECORDED_RESPONSE_PREDICTION",
        "target_semantics": "Native recorded RATA category code; NOT calibrated perceived intensity.",
        "admitted_measurement": "nominal_categorical_response",
        "native_observed_codes": list(range(6)),
        "scale_anchors": None,
        "scale_anchor_status": "NOT_DOCUMENTED_IN_OBTAINED_REPOSITORY",
        "zero_semantics": "OBSERVED_CODE_ZERO; selection/absence/missing meaning not established; do not recode as sensory absence.",
        "permitted_metric": "Multiclass Brier or log loss over six recorded categories; no ordinal-distance or intensity interpretation without additional verified protocol.",
        "panel_provenance": "25 aligned numeric panelist IDs in each original RATA block; expertise/training and sampling protocol not documented in obtained repository.",
        "independence": {
            "participants": 25,
            "source_defined_conditions": 9,
            "records": 225,
            "observed_cells": 2250,
            "traceable_independent_coffee_lots": None,
            "coffee_materials": "One study of Liberica coffee bags; three native roast labels crossed with 0/5/10 leaf-condition codes; raw lot identity and preparation protocol not documented.",
            "cross_coffee_generalization": "NOT_EVALUABLE",
            "duplicate_review": "New source family and 2025 repository, distinct from D0 and R1; no claim that nine treatments are independent beans.",
        },
        "fields": [
            {
                "id": key,
                "source_header": source_label,
                "source_summary_label": label,
                "sheet": "RATA Test",
                "header_row": row,
                "first_data_column": offset + 3,
            }
            for row, offset, source_label, key, label in fields
        ],
        "split_seed": SEED,
        "split_frozen_before_model_fit": True,
        "participant_assignment": {
            p: assigned(p, "liberica_participant") for p in participants
        },
        "condition_assignment": {
            c: assigned(c, "liberica_condition") for c in conditions
        },
        "validation_scopes": [
            "held participant within same study coffee materials",
            "held condition within same study coffee materials",
        ],
        "exclude_from_features": [
            "participant_id",
            "condition_id",
            "coffee_group_id",
            "source_C0",
            "source_C1",
            "hedonic sheets",
            "chemical measurements",
            "effectivity",
            "SPSS recodes",
            "spreadsheet totals and means",
        ],
        "task_masks": {
            "recorded_response": True,
            "sensory_intensity": False,
            "sensory_absence": False,
            "professional_profile": False,
            "individual_perceptual_alignment": False,
            "actual_answer_time": False,
        },
    }
    return {"contract": contract, "records": rows}


def barahona():
    article = etree.fromstring(checked("barahona2020.xml"))
    archive = zipfile.ZipFile(io.BytesIO(checked("barahona2020-supp.zip")))
    docx = zipfile.ZipFile(io.BytesIO(archive.read("FSN3-8-1173-s001.docx")))
    emf = docx.read("word/media/image1.emf")
    # Read native EMR_EXTTEXTOUTW text from the embedded response-form image.
    # The form is used only to verify scale endpoints; never republished.
    offset, strings = 0, []
    while offset + 8 < len(emf):
        kind, size = struct.unpack_from("<II", emf, offset)
        assert size >= 8
        if kind == 84:
            count, start = struct.unpack_from("<II", emf, offset + 44)
            assert start + count * 2 <= size
            strings.append(
                emf[offset + start : offset + start + count * 2].decode("utf-16-le")
            )
        offset += size
    assert {"1", "10", "Lowest", "Highest", "Bitterness", "Sweetness"} <= set(strings)
    source_labels = [
        "Aroma",
        "Acidity",
        "Fragrance",
        "Bitter",
        "Body",
        "Sweet",
        "Residual flavor",
    ]
    keys = [
        "aroma",
        "acidity",
        "fragrance",
        "bitter",
        "body",
        "sweet",
        "residual_flavor",
    ]
    table = article.xpath('//table-wrap[@id="fsn31404-tbl-0004"]/table')[0]
    headers = ["".join(e.itertext()).strip() for e in table.xpath("./thead/tr/*")]
    assert headers == ["Product", *source_labels, "Overall impression"]
    records = []
    for row in table.xpath("./tbody/tr"):
        cells = ["".join(e.itertext()).strip() for e in row]
        assert len(cells) == 9 and re.fullmatch(r"p\d+", cells[0])
        group = "barahona2020:" + cells[0]
        values = dict(zip(keys, map(float, cells[1:8])))
        assert all(1 <= v <= 10 for v in values.values())
        records.append(
            {
                "record_id": group,
                "source_family": "family.barahona_colombian_consumers_2020",
                "group_id": group,
                "sample_id": cells[0],
                "split": assigned(group, "barahona_product"),
                "role": "AUX_COFFEE_WEAK_LABEL",
                "source_C0": "preparation.family.filter_percolation",
                "source_C1": None,
                "attribute_measurements": values,
                "attribute_masks": {key: True for key in keys},
                "attribute_states": {key: "OBSERVED" for key in keys},
                "evidence_unit_id": f"10.1002/fsn3.1404#Table4:{cells[0]}",
            }
        )
    assert len(records) == 18 and len({r["group_id"] for r in records}) == 18
    contract = {
        "source_id": "BARAHONA_COLOMBIAN_CONSUMERS_2020",
        "doi": "10.1002/fsn3.1404",
        "source_url": "https://pmc.ncbi.nlm.nih.gov/articles/PMC7020298/",
        "authors": ["Igor Barahona", "Edis Mauricio Sanmiguel Jaimes", "Jian-Bo Yang"],
        "license": "CC-BY-4.0",
        "rights_scope": "NONCOMMERCIAL_RESEARCH_USE",
        "attribution": "Barahona, I.; Sanmiguel Jaimes, E.M.; Yang, J.-B. (2020). Sensory attributes of coffee beverages and their relation to price and package information. Food Science & Nutrition 8:1173–1186. doi:10.1002/fsn3.1404. Modified: sensory-only Table4 extraction and fixed grouped splits.",
        "raw_sha256": DIGESTS["barahona2020.xml"],
        "supplement_sha256": DIGESTS["barahona2020-supp.zip"],
        "scale_evidence": "Methods 3.1: seven items measure extent of sensory presence; eighth is liking. Appendix A embedded EMF verifies numbered responses1..10 and Lowest/Highest anchors.",
        "scale": {
            "minimum": 1,
            "maximum": 10,
            "low_anchor": "Lowest",
            "high_anchor": "Highest",
            "type": "10-category ordinal, published product means",
        },
        "fields": [
            {"id": key, "source_label": label}
            for key, label in zip(keys, source_labels)
        ],
        "panel_provenance": "130 nontrained Colombian consumers, brief 30-minute explanation/practice; NOT professional panel. Original study 1386 assessments, not publicly obtained.",
        "independence": {
            "source_defined_coffee_product_groups": 18,
            "records": 18,
            "observed_cells": 126,
            "original_individual_assessments_obtained": 0,
            "traceable_independent_coffee_lots": None,
            "unit": "Anonymous brand/product IDs, pooled across cities and participants; underlying lots/batches unverified.",
        },
        "preparation": "Source drip maker:5g ground coffee/10oz water at92C,20mL servings; service delay not documented. C0 filter_percolation follows actual method; C1 source roast degree missing.",
        "target_semantics": "Published consumer mean of ordinal sensory attribute presence, not preference, quality, expert consensus or calibrated mental distance.",
        "admitted_task": "RECORDED_RESPONSE_PREDICTION",
        "permitted_metric": "Source-native ordinal-mean reconstruction MAE, explicitly an encoding error; no cross-source scaling or professional/individual alignment claim.",
        "split_seed": SEED,
        "split_frozen_before_model_fit": True,
        "product_assignment": {r["group_id"]: r["split"] for r in records},
        "exclude_from_features": [
            "sample_id",
            "group_id",
            "source_C0",
            "source_C1",
            "price",
            "package",
            "Overall impression",
            "city",
            "gender",
            "quality",
        ],
        "task_masks": {
            "consumer_ordinal_mean_reconstruction": True,
            "professional_profile": False,
            "individual_perceptual_alignment": False,
            "descriptor_recovery": False,
            "actual_answer_time": False,
        },
        "duplicate_review": "New study family relative to D0/R1. Source product IDs anonymized, lot independence and cross-study brands not verifiable; retain standalone source scope.",
    }
    return {"contract": contract, "records": records}


def croijmans():
    metadata = json.loads(checked("croijmans-dans-api.json"))
    version = metadata["data"]["latestVersion"]
    assert version["license"]["name"] == "CC0-1.0"
    assert (
        "All data in this dataset is published under Creative Commons (CC0"
        in checked("croijmans-readme.txt").decode("cp1252")
    )
    inputs = [
        ("croijmans-coffee-experts.xlsx", "COFFEE_EXPERT", 176238),
        ("croijmans-coffee-wineexperts.xlsx", "WINE_EXPERT", 176235),
        (
            "croijmans-coffee-novices.xlsx",
            "SOURCE_LABELED_NOVICE_COFFEE_CONSUMER",
            176226,
        ),
    ]
    coffee_names = {
        "C1": ["Santa Helena Caturra", "Colombia"],
        "C2": ["Kirimiro Red Bourbon", "Burundi"],
        "C3": ["Knots Family Heirloom varietals", "Ethiopia"],
        "C4": ["Fazenda Rainha Yellow Bourbon", "Brazil"],
        "C5": ["Hacienda Sonora Villa Sarchï", "Costa Rica"],
    }
    grouped = defaultdict(list)
    blank_rows = 0

    def seconds(value):
        return (
            value.hour * 3600
            + value.minute * 60
            + value.second
            + value.microsecond / 1e6
        )

    for filename, population, file_id in inputs:
        source_meta = next(
            f for f in version["files"] if f["dataFile"]["id"] == file_id
        )
        assert not source_meta["restricted"]
        sheet = openpyxl.load_workbook(
            io.BytesIO(checked(filename)), data_only=True
        ).active
        assert sheet.max_column == 18
        assert sheet.cell(1, 1).value == "Participant"
        assert (
            sheet.cell(1, 8).value == "FullAnswer"
            and sheet.cell(1, 12).value == "MainResponse"
        )
        for number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            if row[0] is None:
                assert not any(
                    v is not None for v in row
                ), "Nonempty row lacks participant"
                blank_rows += 1
                continue
            pid, coffee, stimulus_num, modality = row[:4]
            assert (
                isinstance(pid, int)
                and coffee in coffee_names
                and int(coffee[1:]) == stimulus_num
            )
            assert modality in {1, 2} and row[12] in {"S", "A", "E", "O"}
            assert isinstance(row[7], str)
            grouped[(pid, coffee, modality)].append(
                {
                    "source_row": number,
                    "source_file": filename,
                    "population": population,
                    "onset_seconds": seconds(row[4]),
                    "offset_seconds": seconds(row[5]),
                    "duration_seconds": seconds(row[6]),
                    "full_answer_nl": row[7],
                    "source_length_characters": row[8],
                    "source_word_count": row[9],
                    "full_response_nl": row[10],
                    "main_response_nl": row[11],
                    "source_SAE": row[12],
                    "source_response_order": row[13],
                    "source_first_response": row[14],
                    "source_answer_code": row[15],
                    "source_repeated_answer": row[16],
                    "source_double_answer": row[17],
                }
            )
    assert len(grouped) == 627 and sum(map(len, grouped.values())) == 3467
    conflicts = {}
    for (pid, coffee, modality), values in grouped.items():
        signatures = {
            (
                v["full_answer_nl"],
                v["onset_seconds"],
                v["offset_seconds"],
                v["duration_seconds"],
            )
            for v in values
        }
        if len(signatures) > 1:
            conflicts[(pid, coffee)] = (
                "Source modality block mixes distinct FullAnswer and/or timing signatures; entire participant-coffee smell+taste pair excluded without guessing corrections."
            )
    assert set(conflicts) == {(201, "C1"), (110, "C2"), (111, "C3"), (115, "C1")}
    records, quarantine = [], []
    for (pid, coffee, modality), values in sorted(grouped.items()):
        identity = f"croijmans2016:{pid}:{coffee}:{modality}"
        participant = f"croijmans2016:participant:{pid}"
        group = f"croijmans2016:coffee:{coffee}"
        if (pid, coffee) in conflicts:
            quarantine.append(
                {
                    "record_id": identity,
                    "participant_id": participant,
                    "coffee_group_id": group,
                    "source_rows": values,
                    "reason": conflicts[(pid, coffee)],
                }
            )
            continue
        first = values[0]
        onset, offset, duration = (
            first["onset_seconds"],
            first["offset_seconds"],
            first["duration_seconds"],
        )
        timing_valid = duration >= 0 and abs((offset - onset) - duration) < 0.01
        records.append(
            {
                "record_id": identity,
                "source_family": "family.croijmans_majid_flavor_naming_2016",
                "participant_id": participant,
                "source_population": first["population"],
                "coffee_group_id": group,
                "coffee_id": coffee,
                "modality": "SMELL" if modality == 1 else "TASTE",
                "pair_id": f"{participant}|{coffee}",
                "coffee_split": assigned(group, "croijmans_coffee"),
                "participant_split": assigned(participant, "croijmans_participant"),
                "role": (
                    "CORE_PROFESSIONAL"
                    if first["population"] == "COFFEE_EXPERT"
                    else "AUX_COFFEE_WEAK_LABEL"
                ),
                "source_C0": "preparation.family.immersion",
                "source_C1": None,
                "language": "nl",
                "full_answer_nl": first["full_answer_nl"],
                "source_coded_mentions": values,
                "source_unique_main_response_tokens_nl": sorted(
                    {
                        v["main_response_nl"]
                        for v in values
                        if isinstance(v["main_response_nl"], str)
                    }
                ),
                "source_SA_tokens_nl": sorted(
                    {
                        v["main_response_nl"]
                        for v in values
                        if v["source_SAE"] in {"S", "A"}
                        and isinstance(v["main_response_nl"], str)
                    }
                ),
                "recording_onset_seconds": onset,
                "recording_offset_seconds": offset,
                "source_reported_description_seconds": duration,
                "observed_description_seconds": duration if timing_valid else None,
                "timing_mask": timing_valid,
                "full_answer_mask": True,
                "unmentioned_descriptor_state": "NOT_MENTIONED",
                "positive_sensory_truth_mask": False,
                "attribute_intensity_mask": False,
                "actual_product_question_answer_seconds": None,
            }
        )
    assert len(records) == 619 and len(quarantine) == 8
    pairs = defaultdict(dict)
    for record in records:
        pairs[record["pair_id"]][record["modality"]] = record["record_id"]
    complete_pairs = [
        {
            "pair_id": key,
            "smell_record_id": values["SMELL"],
            "taste_record_id": values["TASTE"],
        }
        for key, values in sorted(pairs.items())
        if set(values) == {"SMELL", "TASTE"}
    ]
    participants = sorted({r["participant_id"] for r in records})
    contract = {
        "source_id": "CROIJMANS_MAJID_COFFEE_LANGUAGE_2016",
        "doi": "10.17026/dans-zke-2wgq",
        "article_doi": "10.1371/journal.pone.0155845",
        "source_url": "https://ssh.datastations.nl/dataset.xhtml?persistentId=doi:10.17026/DANS-ZKE-2WGQ",
        "license": "CC0-1.0",
        "rights_scope": "NONCOMMERCIAL_RESEARCH_USE",
        "attribution": "Majid, A.; Croijmans, I. (2016). Human olfaction at the intersection of language, culture and biology. DANS. doi:10.17026/dans-zke-2wgq. Article: Croijmans & Majid, PLoS ONE11(6):e0155845. Modified: coffee-only extraction, grouped source-code QA and split assignments; Dutch text preserved.",
        "license_evidence": "DANS latestVersion2.1 CC0-1.0 plus original README explicit CC0 for all dataset data; selected files unrestricted.",
        "files": [
            {"filename": filename, "file_id": file_id, "sha256": DIGESTS[filename]}
            for filename, _, file_id in inputs
        ],
        "coffee_identity": {
            key: {"name": value[0], "country": value[1]}
            for key, value in coffee_names.items()
        },
        "grouping": "Five distinct single-estate coffee origins; same roasting procedure in one batch is shared condition, not proof of one original green-bean lot. All participants and modalities for one coffee remain together in coffee-holdout.",
        "preparation": "Original methods: five coffees roasted together with same procedure, sealed100g lots; medium-fine grinding within3h, freshly cupped following SCAA guidelines. Source-native C1 degree not reported. Immersion is sourced to cupping, not automatic seven-bin mapping.",
        "panel_provenance": "20 qualified coffee baristas/roasters/brokers;22 wine experts;21 source-labeled novices who consumed coffee and wine at least weekly. Native Dutch descriptions, untrained consumers not equated to complete coffee beginners or target-user population.",
        "reference_semantics": "Independent observed professional expressions or independent consumer expressions, not calibrated professional intensity or unique coffee truth. Original paper explicitly says no correct answer for coffee/wine descriptions.",
        "admitted_tasks": [
            "OBSERVED_DESCRIPTOR_RECOVERY",
            "RECORDED_RESPONSE_PREDICTION",
        ],
        "professional_expression_alignment": "Possible only after evaluator fixes disjoint reference participants; label as alignment with recorded expert language, not normative true sensory profile.",
        "source_coding_limitations": "MainResponse is linguistic coding and may include negation/comparison from FullResponse. Source S/A classes are source-based/non-source-based language, not signed positive sensory labels. Preserve raw clauses; do not turn tokens or omissions into sensory presence/absence.",
        "source_SAE": {
            "S": "source-based term",
            "A": "non-source-based term",
            "E": "evaluative term",
            "O": "other",
        },
        "timing": "README defines Onset/Offset from audio recording and Time as rough whole-description duration. Repeated term rows duplicate this interval. Count duration once per valid full description; NOT term-specific time, product question-answer time, or evidence of system time savings.",
        "independence": {
            "source_defined_single_estate_coffees": 5,
            "raw_participants": 63,
            "raw_coded_term_rows": 3467,
            "raw_response_groups": 627,
            "quarantined_participant_coffee_pairs": 4,
            "quarantined_response_groups": 8,
            "admitted_response_groups": 619,
            "complete_smell_taste_pairs": len(complete_pairs),
            "original_physical_lot_ids": None,
        },
        "split_seed": SEED,
        "split_frozen_before_model_fit": True,
        "coffee_assignment": {
            f"croijmans2016:coffee:{c}": assigned(
                f"croijmans2016:coffee:{c}", "croijmans_coffee"
            )
            for c in coffee_names
        },
        "participant_assignment": {
            p: assigned(p, "croijmans_participant") for p in participants
        },
        "validation_scopes": [
            "held coffee across all people/modalities",
            "held participant within these same five coffees; separate scope",
        ],
        "task_masks": {
            "native_recorded_language": True,
            "paired_independent_expert_consumer_expressions": True,
            "whole_description_recording_time": True,
            "actual_product_answer_time": False,
            "professional_intensity": False,
            "positive_sensory_truth": False,
            "individual_perceptual_alignment": False,
        },
        "exclude_from_features": [
            "participant_id",
            "coffee_id",
            "coffee_group_id",
            "pair_id",
            "source_population",
            "origin/name",
            "recording timestamps",
            "reference participant outputs",
            "SAE evaluative/other classes when used as sensory-reference outputs",
        ],
        "duplicate_review": "New DANS2016 source family, distinct from R1 and D0 provenance. Anonymous observations preserve source IDs; no extra group counted per expert, descriptor row, smell/taste modality or public CSV mirror.",
    }
    return {
        "contract": contract,
        "records": records,
        "complete_pairs": complete_pairs,
        "quarantine": quarantine,
        "blank_spreadsheet_rows_ignored": blank_rows,
    }


def croijmans_sorting():
    """Independent participant perceptual sorting; never reconstructed coordinates."""
    workbook = openpyxl.load_workbook(
        io.BytesIO(checked("croijmans-stimuli-sorting.xlsx")), data_only=True
    )
    sheet = workbook["Stimulus Comparisons"]
    mirror = list(
        csv.reader(
            io.StringIO(checked("croijmans-stimulus-comparisons.tab").decode()),
            delimiter="\t",
        )
    )
    original = list(sheet.values)
    assert len(original) == len(mirror) == 21
    assert len(original[0]) == len(mirror[0]) == 23
    assert [str(v) for v in original[0]] == mirror[0]
    characteristics = {
        int(row[0]): row
        for row in list(workbook["ParticipantCharacteristics"].values)[1:]
    }
    assert set(characteristics) == set(range(401, 421))
    coffee_pairs = list(itertools.combinations(range(1, 6), 2))
    participants, records = {}, []
    for row_number, (row, mirror_row) in enumerate(zip(original[1:], mirror[1:]), 2):
        participant_number = int(row[0])
        assert participant_number == int(mirror_row[0])
        participant_id = f"croijmans2016:sorting_participant:{participant_number}"
        participant_split = assigned(participant_id, "croijmans_sorting_participant")
        order_code = int(characteristics[participant_number][4])
        assert order_code in (1, 2)
        # Only source IDs/order are kept; demographics have no task role.
        participants[participant_id] = {
            "split": participant_split,
            "source_order_code": order_code,
        }
        values = []
        for column, (a, b) in enumerate(coffee_pairs, 4):
            header = original[0][column - 1]
            assert tuple(map(int, re.findall(r"\d+", header))) == (a, b)
            raw_value = row[column - 1]
            assert str(raw_value) == mirror_row[column - 1]
            value = float(str(raw_value).replace(",", "."))
            assert math.isfinite(value) and value >= 0
            values.append(value)
            coffee_ids = [f"croijmans2016:coffee:C{v}" for v in (a, b)]
            coffee_splits = [assigned(v, "croijmans_coffee") for v in coffee_ids]
            # Any edge touching a reserved coffee is held out. No edge involving
            # a confirmation coffee can leak into a training distance objective.
            coffee_pair_split = (
                "DEVELOPMENT"
                if set(coffee_splits) == {"DEVELOPMENT"}
                else "CONFIRMATION"
            )
            records.append(
                {
                    "record_id": f"{participant_id}|C{a}|C{b}",
                    "source_family": "family.croijmans_majid_flavor_naming_2016",
                    "participant_id": participant_id,
                    "participant_split": participant_split,
                    "coffee_a_id": coffee_ids[0],
                    "coffee_b_id": coffee_ids[1],
                    "coffee_a_split": coffee_splits[0],
                    "coffee_b_split": coffee_splits[1],
                    "coffee_pair_split": coffee_pair_split,
                    "distance_mm": value,
                    "distance_mask": True,
                    "distance_state": "OBSERVED",
                    "source_cell": f"Stimulus Comparisons!{sheet.cell(row_number, column).coordinate}",
                    "source_value": raw_value,
                    "role": "AUX_INDEPENDENT_PERCEPTUAL_SORTING",
                    "modality": "FLAVOR_SIMILARITY_MODALITY_UNSPECIFIED",
                    "source_C0": None,
                    "source_C1": None,
                    "actual_product_answer_seconds": None,
                }
            )
        source_mean = float(str(row[1]).replace(",", "."))
        assert (
            abs(sum(values) / len(values) - source_mean) <= 0.0051
        ), "Source mean disagrees with original distances"
    assert len(records) == 200 and len({r["record_id"] for r in records}) == 200
    assert len(participants) == 20 and all(r["distance_mask"] for r in records)
    assert Counter(p["source_order_code"] for p in participants.values()) == {
        1: 10,
        2: 10,
    }
    coffee_assignment = {
        f"croijmans2016:coffee:C{i}": assigned(
            f"croijmans2016:coffee:C{i}", "croijmans_coffee"
        )
        for i in range(1, 6)
    }
    contract = {
        "source_id": "CROIJMANS_MAJID_INDEPENDENT_SORTING_2016",
        "doi": "10.17026/dans-zke-2wgq",
        "article_doi": "10.1371/journal.pone.0155845",
        "source_url": "https://ssh.datastations.nl/dataset.xhtml?persistentId=doi:10.17026/DANS-ZKE-2WGQ",
        "license": "CC0-1.0",
        "rights_scope": "NONCOMMERCIAL_RESEARCH_USE",
        "attribution": "Majid, A.; Croijmans, I. (2016), DANS doi:10.17026/dans-zke-2wgq; Croijmans & Majid (2016), PLoS ONE doi:10.1371/journal.pone.0155845. Modified: coffee-only original distance extraction, decimal-comma parsing and frozen grouped splits.",
        "license_evidence": "Same unrestricted DANS deposit and CC0 README as language data; independent sorting is another measurement in this source, not another source or five additional coffees.",
        "files": [
            {"filename": name, "file_id": file_id, "sha256": DIGESTS[name]}
            for name, file_id in [
                ("croijmans-stimuli-sorting.xlsx", 176227),
                ("croijmans-stimulus-comparisons.tab", 176298),
            ]
        ],
        "protocol": "Original article Methods/Comparability and README: a separate group of20 naive participants sorted the five coffees and five wines by perceived similarity, placing glasses on paper; half coffee first. Recorded x/y coordinates were transformed by the authors into Euclidean inter-stimulus distances. Original workbook contains distances, not x/y coordinates. The sorting subsection does not prescribe smell-only versus sip/taste-only sampling; do not invent that detail.",
        "measurement": "Source-native participant-specific spatial dissimilarity in millimeters; low means glasses were placed closer. Not descriptor intensity, ordinal CATA code, calibrated sensory metric, or physical coffee distance.",
        "independence": {
            "independent_sorting_participants": 20,
            "coffee_pairs_per_participant": 10,
            "observed_pair_distances": 200,
            "same_five_coffees_as_expression_study": True,
            "new_coffee_groups": 0,
            "new_source_families": 0,
            "raw_coordinates_available": False,
            "original_physical_lot_ids": None,
        },
        "missing_protocol": "All200 pair distances observed. No unrecorded pair, descriptor, answer time or coordinate is generated. Zero would be a valid observed spatial distance, not absence of sensation.",
        "qa": "All200 original XLSX distance strings equal the repository's TSV mirror; every participant's mean matches the published source mean within its two-decimal rounding precision. Source demographics, wine distances and summary means excluded from model records.",
        "split_seed": SEED,
        "split_frozen_before_model_fit": True,
        "coffee_assignment": coffee_assignment,
        "participant_assignment": {
            p: value["split"] for p, value in participants.items()
        },
        "participant_order_codes": {
            p: value["source_order_code"] for p, value in participants.items()
        },
        "pair_split_rule": "Reuse previously frozen language coffee split: C1/C2/C5 development, C3/C4 confirmation. Training distance edges must have both endpoints in development; any edge touching a confirmation coffee is held. Also freeze independent sorting participants before any model or metric; participant scope differs from coffee scope.",
        "admitted_task": "INDEPENDENT_RECORDED_PERCEPTUAL_GEOMETRY_REFERENCE",
        "task_masks": {
            "source_native_perceptual_dissimilarity": True,
            "professional_profile_alignment": False,
            "individual_perceptual_alignment_of_this_system": False,
            "observed_descriptor_recovery": False,
            "actual_product_answer_time": False,
        },
        "evaluation_limit": "May evaluate a predeclared coffee representation against an independent participant's observed spatial sorting. This is not a trial in which those participants used this system, and cannot establish personalized product alignment or time savings. Five coffees give only10 distinct stimulus pairs;200 observations are not200 independent coffees or pairs.",
        "exclude_from_features": [
            "sorting participant_id",
            "coffee identity/origin",
            "source group",
            "source task order",
            "held-out pair distance",
            "wine distances",
            "age/gender",
            "summary means",
            "nonexistent x/y coordinates",
        ],
    }
    return {"contract": contract, "records": records}


def increment_manifest(ended_utc=None):
    """Metadata-only receipt; raw source rows and all models remain private."""
    routes = [
        (
            "LIBERICA_RATA",
            "10.17632/m3n2gc4dv6.1",
            "https://data.mendeley.com/datasets/m3n2gc4dv6/1",
            ["liberica-data.zip", "liberica-mendeley.html"],
            "CC-BY-4.0",
            "ADMITTED",
            "Original RATA respondent blocks225x10 parsed. Native0..5 codes have no verified anchor definitions; nominal recorded-response only. Nine treatments are not nine independent coffees.",
        ),
        (
            "BARAHONA_PRODUCTS",
            "10.1002/fsn3.1404",
            "https://pmc.ncbi.nlm.nih.gov/articles/PMC7020298/",
            ["barahona2020.xml", "barahona2020-supp.zip"],
            "CC-BY-4.0",
            "ADMITTED",
            "Original Table4 provides18 product meansx7 sensory columns. Original form embedded EMF verifies1..10 Lowest/Highest response scale. Overall impression/liking, price and quality excluded; individual respondent matrix not obtained.",
        ),
        (
            "CROIJMANS_PAIRED_LANGUAGE_AND_SORTING",
            "10.17026/dans-zke-2wgq",
            "https://ssh.datastations.nl/dataset.xhtml?persistentId=doi:10.17026/DANS-ZKE-2WGQ",
            [
                "croijmans-dans-api.json",
                "croijmans-readme.txt",
                "croijmans2016.xml",
                "croijmans-methods.pdf",
                "croijmans-coffee-experts.xlsx",
                "croijmans-coffee-wineexperts.xlsx",
                "croijmans-coffee-novices.xlsx",
                "croijmans-stimuli-sorting.xlsx",
                "croijmans-stimulus-comparisons.tab",
            ],
            "CC0-1.0",
            "ADMITTED",
            "Five coffees paired with independent coffee-expert/wine-expert/novice recorded language, plus separate20-person similarity sorting.8 conflicting response groups quarantined, not repaired; same source/coffee count for both measurements.",
        ),
        (
            "CATA_INTENSITY_METHOD",
            "10.1111/joss.12833",
            "https://www.colibri.udelar.edu.uy/jspui/handle/20.500.12008/41076",
            ["jaeger12833.pdf", "jaeger12833-record.html"],
            "CC-BY-4.0",
            "METHOD_EVIDENCE_ONLY",
            "Original full paper body read. Seven online pictured-food studies, not coffee tasting. CATA selection probability and expected intensity are distinct; raw responses available only by author request. No requests sent.",
        ),
        (
            "RTD_PROFILING_METHOD",
            "10.1111/joss.12839",
            "https://www.researchgate.net/publication/370609890_Comparison_of_conventional_and_consumer-based_sensory_profiling_methods_for_ready-to-drink_coffee_beverages",
            ["yoon12839-crossref.json", "yoon12839-openalex.json"],
            "STANDARD_PUBLISHER_COPYRIGHT_NO_RAW_REUSE_LICENSE_VERIFIED",
            "METHOD_EVIDENCE_ONLY",
            "Original author-uploaded paper methods/results/tables/limitations read through web text; publisher closed, no local original full-PDF hash. Six milk-based RTD products;10 DA panelists and separate70-person CATA/FP/FL groups. Raw matrix only by request; group configuration agreement is not individual truth, protocol-level duration not product-question duration.",
        ),
        (
            "COLD_BREW_CONSUMER_2019",
            "10.3390/foods8080344",
            "https://pmc.ncbi.nlm.nih.gov/articles/PMC6723667/",
            ["heo2019.xml", "heo2019-supp.zip"],
            "CC-BY-4.0",
            "NOT_ADMITTED_LIMITED_AGGREGATE_ONLY",
            "Actual supplement contains figures, no participant CATA matrix. Article Table3 has3 intensity-code means plus liking; source prose13 samples versus14 table rows includes repeated commercial control. Four bean materials have paired brewing conditions, but cold storage and aggregation limit additional value; aggregate adapter deferred after stronger paired source obtained, not claimed inaccessible.",
        ),
        (
            "EEG_DESCRIPTIVE_PANEL_2025",
            "10.3389/fnhum.2025.1661214",
            "https://pmc.ncbi.nlm.nih.gov/articles/PMC12546097/",
            ["eeg2025.xml", "eeg2025-supp.zip"],
            "CC-BY-4.0_ARTICLE",
            "NO_ADMISSIBLE_ROW_MATRIX",
            "Actual supplementary archive contains images, not coffee-panel response matrix; raw data by request. No EEG/liking proxies become sensory labels.",
        ),
        (
            "COFFEE_CUALITY_2026",
            "10.3390/foods15040678",
            "https://doi.org/10.3390/foods15040678",
            ["cuality2026.xml", "cuality2026-supp.zip"],
            "CC-BY-4.0_ARTICLE_THIRD_PARTY_FORMS_NOT_IMPORTED",
            "NO_ADMISSIBLE_ROW_MATRIX",
            "Actual nested supplement PDF contains scorecards, not response rows.56 professionals/12 coffees are study counts only, not acquired observations. Quality/JAR scores not sensory intensity; no proprietary form republished.",
        ),
        (
            "CROSS_CULTURE_ESPRESSO_2025",
            "10.3390/foods14040593",
            "https://doi.org/10.3390/foods14040593",
            ["crossculture2025.xml", "crossculture2025-supp.zip"],
            "CC-BY-4.0_ARTICLE_DATA_RESTRICTED",
            "RESTRICTED_DATA_NOT_ACQUIRED",
            "Actual supplement is roast table only. Article explicitly restricts raw dataset for ethics/privacy.46 experts and8 coffees not acquired; no additional requests or circumvention.",
        ),
        (
            "CROSS_BRAND_TEMPORAL_LIKING_2025",
            "10.3390/app15020948",
            "https://doi.org/10.3390/app15020948",
            ["crossbrand2025.xml"],
            "CC-BY-4.0_ARTICLE",
            "NO_ADMISSIBLE_ROW_MATRIX",
            "Four RTD coffees, temporal liking is separate target. Dataset upon request subject to intended-use approval, no raw matrix obtained or requests sent.",
        ),
        (
            "WET_PROCESSING_PANEL_2019",
            "10.3389/fmicb.2019.02621",
            "https://pmc.ncbi.nlm.nih.gov/articles/PMC6863779/",
            ["processing2019.xml", "processing2019-supp.zip"],
            "CC-BY-4.0",
            "NO_ADMISSIBLE_ROW_MATRIX",
            "Trained panel27 sensory attributes in article; actual16-page supplement contains genetic/processing tables and plots, no joint sensory rows. No matrix digitized or fabricated from PCA.",
        ),
        (
            "REMOTE_TESTING_2021",
            "10.1016/j.foodqual.2021.104437",
            "https://pmc.ncbi.nlm.nih.gov/articles/PMC8548442/",
            ["remote2021.xml", "remote2021-supp.zip"],
            "ELSEVIER_COVID_RESOURCE_CONDITIONAL_RESEARCH_PERMISSION_NOT_ASSUMED_CC",
            "NO_ADMISSIBLE_ROW_MATRIX",
            "Original full-text protocol and actual supplement inspected. Three mocha brands across historical batches/lab versus remote; DOCX supplement is ANOVA for other foods, not individual coffee matrix. Conditional reuse not needed because no data admitted.",
        ),
    ]
    download_urls = {
        "croijmans-dans-api.json": "https://ssh.datastations.nl/api/datasets/:persistentId/?persistentId=doi:10.17026/DANS-ZKE-2WGQ"
    }
    for receipt in SOURCES.glob("*_download_receipt.private.json"):
        for item in json.loads(receipt.read_text()):
            if "url" in item:
                download_urls[item["file"]] = item["url"]
    route_records = []
    for route_id, doi, url, files, rights, status, outcome in routes:
        file_records = []
        for name in files:
            data = (SOURCES / name).read_bytes()
            file_records.append(
                {
                    "private_filename": name,
                    "sha256": digest(data),
                    "bytes": len(data),
                    "download_url": download_urls.get(name),
                }
            )
        route_records.append(
            {
                "route_id": route_id,
                "doi": doi,
                "source_url": url,
                "new_direction_vs_R1": True,
                "license_scope": rights,
                "result": status,
                "actual_outcome_and_model_use": outcome,
                "actual_files": file_records,
            }
        )
    names = [
        "liberica_rata_records.private.json",
        "barahona_ordinal_means.private.json",
        "croijmans_paired_responses.private.json",
        "croijmans_sorting_distances.private.json",
    ]
    adapters, source_data = [], {}
    for name in names:
        raw = (PRIVATE / name).read_bytes()
        data = json.loads(raw)
        source_data[name] = data
        contract = data["contract"]
        adapters.append(
            {
                "source_id": contract["source_id"],
                "path": "<OWNER_V2>/revisions/r2/" + name,
                "sha256": digest(raw),
                "bytes": len(raw),
                "admitted_records": len(data["records"]),
                "license": contract["license"],
                "attribution": contract["attribution"],
                "measurement_task": contract.get(
                    "admitted_task", contract.get("admitted_tasks")
                ),
                "measurement_semantics": contract.get(
                    "target_semantics",
                    contract.get("measurement", contract.get("reference_semantics")),
                ),
                "independence": contract["independence"],
                "split_frozen_before_model_fit": True,
                "split_counts": {
                    key: dict(Counter(contract[key].values()))
                    for key in [
                        "participant_assignment",
                        "condition_assignment",
                        "product_assignment",
                        "coffee_assignment",
                    ]
                    if key in contract
                },
                "task_masks": contract.get("task_masks", {}),
                "actual_fields": contract.get(
                    "fields",
                    "See private source contract; raw expressions and coded vocabulary remain private.",
                ),
            }
        )
    expression = source_data["croijmans_paired_responses.private.json"]
    first_download = min((SOURCES / item[3][0]).stat().st_mtime for item in routes)
    started = datetime.fromtimestamp(first_download, timezone.utc)
    ended = (
        datetime.fromisoformat(ended_utc.replace("Z", "+00:00")) if ended_utc else None
    )
    return {
        "schema_version": "M2_R2_DATA_INCREMENT_1",
        "base_checkpoint": "8c41e89214229c56849f04efaf61162c38b757d2",
        "source_work_block": {
            "started_utc": started.isoformat(),
            "timestamp_basis": "Exact earliest acquired source-file timestamp; preparatory reading/search began earlier, approximately13:15UTC and is not included in measured acquisition elapsed time.",
            "ended_utc": ended.isoformat() if ended else None,
            "elapsed_minutes": (
                round((ended - started).total_seconds() / 60, 2) if ended else None
            ),
            "status": "COMPLETE" if ended else "IN_PROGRESS",
            "stop_reason": (
                "Completed three new licensed data sources including actual multi-coffee paired language and independent perceptual distances; closed already-open routes after inspecting their actual attachments. No idle wait or quota-driven unusable samples."
                if ended
                else None
            ),
        },
        "private_owner_root": "<OWNER_V2>/revisions/r2",
        "public_scope": "Only source/rights/count/mask/split receipt and acquisition adapter; no source row matrix, free text, participant data, weights, model outputs or proprietary forms copied to public Git.",
        "acquisition_routes": route_records,
        "adapters": adapters,
        "actual_increment": {
            "new_admitted_source_families": 3,
            "admitted_measurement_adapters": 4,
            "source_defined_coffee_product_groups": 23,
            "single_estate_coffee_identities": 5,
            "anonymous_product_groups": 18,
            "traceable_independent_green_lots": None,
            "liberica_material_lots": "UNDOCUMENTED_NOT_COUNTED_AS_NINE_COFFEES",
            "liberica_treatment_conditions": 9,
            "liberica_participant_condition_records": 225,
            "liberica_observed_categorical_cells": 2250,
            "consumer_product_mean_records": 18,
            "consumer_product_mean_cells": 126,
            "raw_expression_groups": 627,
            "design_expression_groups_if_complete_63x5x2": 630,
            "unrecorded_design_expression_groups": 3,
            "admitted_expression_groups": 619,
            "quarantined_expression_groups": 8,
            "quarantined_participant_coffee_pairs": 4,
            "complete_clean_smell_taste_pairs": 308,
            "clean_unpaired_expression_groups": 3,
            "clean_expression_recording_durations": sum(
                r["timing_mask"] for r in expression["records"]
            ),
            "inconsistent_expression_timing_records": sum(
                not r["timing_mask"] for r in expression["records"]
            ),
            "independent_sorting_participants": 20,
            "sorting_distances": 200,
            "distinct_sorting_coffee_pairs": 10,
            "additional_coffees_from_sorting": 0,
            "new_complete_production_C0_C1_joint_groups": 0,
            "actual_product_question_response_times": 0,
            "professional_source_expressions": sum(
                r["source_population"] == "COFFEE_EXPERT" for r in expression["records"]
            ),
            "new_professional_intensity_profile_groups": 0,
        },
        "target_gap": {
            "requested_source_product_or_useful_case_range": [30, 60],
            "source_product_groups_obtained": 23,
            "product_group_shortfall_to_30": 7,
            "multi_coffee_paired_source_obtained": True,
            "useful_paired_expression_cases_obtained": 308,
            "paired_cases_are_independent_coffees": False,
            "strict_traceable_lot_target_status": "NOT_VERIFIED; product/condition/participant counts cannot prove independent green lots.",
            "quota_status": "Nonblocking goal: multi-coffee paired source and useful cases obtained;30 independent products not reached. Do not sum unlike response units into a coffee count.",
        },
        "global_masks_and_limits": [
            "Unmentioned is NOT_MENTIONED, never absent or zero.",
            "Native intensity codes, citation, spatial distance, quality, liking and free expression remain separate units/tasks.",
            "Croijmans coded MainResponse may contain negation/comparison; native token occurrence is not signed positive sensory truth.",
            "Liberica original0..5 are nominal response categories while anchors remain undocumented; no interval/ordinal assumption.",
            "Barahona means retain original1..10 source coding; no professional or individual perception claim.",
            "Sorting references come from separate people, but they did not use this system; no claim of system individual perceptual alignment.",
            "All new participant/product/coffee splits were written before model evaluation. Previously viewed R1/D0 confirmations remain historical.",
            "Native roast labels are not automatically converted into the production seven-bin taxonomy; missing C0/C1 remain null.",
            "Whole-expression recording durations are not product question times and are forbidden router features.",
        ],
        "parser": {
            "path": "db/scripts/acquire_m2_r2.py",
            "sha256": digest(Path(__file__).read_bytes()),
            "runtime": "Python3 with openpyxl and lxml",
            "verification": "All four source adapters replayed byte-identically; Python compilation passed.",
            "replay": [
                "python db/scripts/acquire_m2_r2.py --source " + source
                for source in ["liberica", "barahona", "croijmans", "croijmans_sorting"]
            ],
            "assertions": [
                "Frozen original source SHA256; no silent refresh",
                "Frozen adapter bytes on replay",
                "Participant and condition alignment across all10 Liberica original blocks",
                "Source form scale anchors verified before Barahona parsing",
                "Source modality/full-answer conflicts quarantined at entire participant-coffee pair",
                "Sorting200 original cells identical to independent repository TSV mirror and source mean rounding",
                "Complete native numeric response masks, missing not silently zero",
                "C0/C1 and forbidden task semantics explicit",
            ],
        },
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--source", choices=["liberica", "barahona", "croijmans", "croijmans_sorting"]
    )
    parser.add_argument("--manifest", action="store_true")
    parser.add_argument(
        "--ended-utc",
        help="Exact UTC acquisition finish for the final metadata-only manifest",
    )
    args = parser.parse_args()
    if args.manifest:
        print(
            json.dumps(
                save(
                    PUBLIC / "data_increment_manifest.json",
                    increment_manifest(args.ended_utc),
                ),
                indent=2,
            )
        )
        return
    parser.error("--source or --manifest required") if args.source is None else None
    value = {
        "liberica": liberica,
        "barahona": barahona,
        "croijmans": croijmans,
        "croijmans_sorting": croijmans_sorting,
    }[args.source]()
    name = {
        "liberica": "liberica_rata_records.private.json",
        "barahona": "barahona_ordinal_means.private.json",
        "croijmans": "croijmans_paired_responses.private.json",
        "croijmans_sorting": "croijmans_sorting_distances.private.json",
    }[args.source]
    result = save(PRIVATE / name, value, freeze=True)
    splits = {
        key: dict(Counter(value["contract"][key].values()))
        for key in [
            "participant_assignment",
            "condition_assignment",
            "product_assignment",
            "coffee_assignment",
        ]
        if key in value["contract"]
    }
    print(
        json.dumps(
            {**result, "records": len(value["records"]), "splits": splits}, indent=2
        )
    )


if __name__ == "__main__":
    main()

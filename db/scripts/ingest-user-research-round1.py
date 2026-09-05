#!/usr/bin/env python3
"""Restricted source extraction and public coded survey receipt (no training role).

Set COFFEE_FLAVOR_USER_RESEARCH_ROOT to an owner-controlled, mode-700 root
containing inventory.json and sources. Numbers must have a native Excel export
in that root; the original Numbers payload remains the audited source.
"""
from __future__ import annotations

import csv
import hashlib
import json
import os
from collections import Counter, defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "db/data/user-research-round1"


def digest(data):
    return hashlib.sha256(data).hexdigest()


def write_json(path, value):
    path.write_text(json.dumps(value, ensure_ascii=False, sort_keys=True, indent=2) + "\n")


def table(name, fields, rows):
    with (OUT / name).open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fields, delimiter="\t", lineterminator="\n")
        w.writeheader()
        w.writerows(rows)


def generate():
    from openpyxl import load_workbook

    restricted = Path(os.environ["COFFEE_FLAVOR_USER_RESEARCH_ROOT"])
    if restricted.stat().st_mode & 0o077:
        raise ValueError("Restricted root must not be accessible to group/other users")
    inv = json.loads((restricted / "inventory.json").read_text())
    if len(inv) != 11:
        raise ValueError("Expected eleven submitted source files; do not silently exclude inputs")
    clusters = defaultdict(list)
    for item in inv:
        original = Path(item["restricted_path"])
        actual = digest(original.read_bytes())
        if actual != item["sha256"]:
            raise ValueError("Source changed since intake: " + item["file_id"])
        clusters[actual].append(item["file_id"])
    duplicate = [ids for ids in clusters.values() if len(ids) > 1]
    if duplicate != [["F003", "F004"]] or len(clusters) != 10:
        raise ValueError("ROUND3O_USER_RESEARCH_INGESTION_CONFLICT: duplicate audit differs")

    OUT.mkdir(parents=True, exist_ok=True)
    raw, audit, ledger, qualitative, anomalies = [], [], [], [], []
    definitions = {}
    note_codes = {
        ("F001", "Q13", "primary"): ["WHITE_FLORAL_NEEDS_EXPLANATION", "ORANGE_BLOSSOM_NEEDS_EXPLANATION", "ROSE_NEEDS_EXPLANATION"],
        ("F001", "OVERALL", "note"): ["NONLITERAL_FEEDBACK_NO_POLICY_INFERENCE"],
        ("F005", "Q2", "note"): ["ADDITIONAL_DRINK_CATEGORY"],
        ("F009", "Q13", "note"): ["ORANGE_BLOSSOM_NEEDS_EXPLANATION"],
        ("F009", "Q16", "note"): ["QUESTION_BUDGET_HAS_UPPER_BOUND"],
        ("F009", "Q17", "note"): ["UNKNOWN_AND_INDISTINGUISHABLE_DIFFER"],
        ("F009", "OVERALL", "note"): ["QUESTION_BURDEN", "DESIRE_FOR_AFFIRMATION"],
        ("F010", "OVERALL", "note"): ["VALUE_PROPOSITION_UNCLEAR"],
    }

    def anomaly(fid, qid, kind, disposition):
        anomalies.append(dict(anomaly_id=f"AN{len(anomalies)+1:03d}", file_id=fid,
                              question_id=qid, anomaly_type=kind, disposition=disposition))

    for item in inv:
        fid = item["file_id"]
        included = fid == min(clusters[item["sha256"]])
        source = Path(item["restricted_path"])
        if item["format"] == "numbers":
            source = restricted / (fid + "-export.xlsx")
            if not source.exists():
                raise ValueError("Numbers native export required; input cannot be excluded")
            anomaly(fid, "Q2", "NUMBERS_READER_LOST_CELLS", "NATIVE_NUMBERS_EXPORT_RECOVERS_ALL_TWENTY_RESPONSES")
        ws = load_workbook(source, data_only=True)["用户问卷"]
        entries = []
        for row in ws.iter_rows():
            values = [cell.value for cell in row]
            if isinstance(values[0], (int, float)) and 1 <= values[0] <= 20 and values[3]:
                qid = "Q" + str(int(values[0]))
                entries.append(dict(question_id=qid, primary=values[8], note=values[9],
                                    primary_locator=f"{ws.title}!I{row[0].row}",
                                    note_locator=f"{ws.title}!J{row[0].row}"))
                definitions.setdefault(qid, dict(question=values[3], options=values[4:8]))
        if [e["question_id"] for e in entries] != [f"Q{q}" for q in range(1, 21)]:
            raise ValueError("Incomplete question extraction: " + fid)
        overall = []
        for row in ws.iter_rows(min_row=41):
            for c in row:
                if c.value is not None:
                    value = str(c.value).replace("可选：你最想让这个产品改进的一点是：", "").strip()
                    if value:
                        overall.append(dict(value=value, locator=f"{ws.title}!{c.coordinate}"))
        raw.append(dict(**item, entries=entries, overall=overall, primary_included=included,
                        extraction_payload_sha256=digest(source.read_bytes())))
        audit.append(dict(file_id=fid, participant_id=item["participant_id"], file_sha256=item["sha256"],
                          file_format=item["format"], duplicate_cluster_id="D001" if len(clusters[item["sha256"]])>1 else "",
                          primary_included=str(included).lower(),
                          restricted_pointer=f"restricted://user-research-round1/{fid}",
                          extraction_method="NUMBERS_NATIVE_EXPORT_OPENPYXL" if item["format"]=="numbers" else "OPENPYXL",
                          extraction_payload_sha256=digest(source.read_bytes())))
        for entry in entries + [dict(question_id="OVERALL", primary="", note=o["value"]) for o in overall]:
            qid = entry["question_id"]
            primary = str(entry["primary"] or "").strip()
            note = str(entry["note"] or "").strip()
            code = primary.upper() if primary.upper() in ("A", "B", "C", "D") else "OPEN" if primary else "MISSING"
            supplemental = note.upper() if note.upper() in ("A", "B", "C", "D") else ""
            if qid != "OVERALL":
                ledger.append(dict(file_id=fid, participant_id=item["participant_id"], question_id=qid,
                                   coded_answer=code, supplementary_code=supplemental,
                                   primary_included=str(included).lower(),
                                   open_response_ref=digest(primary.encode()) if code=="OPEN" else "",
                                   restricted_pointer=f"restricted://user-research-round1/{fid}/{qid}"))
            for field, text in (("primary", primary if code=="OPEN" else ""), ("note", note)):
                if not text:
                    continue
                if text in ("N/A", "NA"):
                    anomaly(fid, qid, "NONRESPONSE_MARKER", "RETAIN_RESTRICTED_EXCLUDE_QUALITATIVE")
                    continue
                if text.isdigit() and len(text)==1:
                    anomaly(fid, qid, "SINGLE_DIGIT_FORMAT_ARTIFACT_CANDIDATE", "RETAIN_RESTRICTED_EXCLUDE_QUALITATIVE")
                    continue
                if supplemental:
                    anomaly(fid, qid, "SUPPLEMENTARY_LETTER", "RETAIN_SEPARATELY_DO_NOT_OVERRIDE_PRIMARY")
                if field=="primary":
                    anomaly(fid, qid, "OPEN_PRIMARY_RESPONSE", "QUALITATIVE_MULTILABEL_NOT_ABCD")
                for category in note_codes.get((fid, qid, field), ["UNINTERPRETED_REQUIRES_OWNER_REVIEW"]):
                    qualitative.append(dict(qualitative_id=f"QL{len(qualitative)+1:03d}", file_id=fid,
                                            question_id=qid, source_field=field, qualitative_category=category,
                                            response_sha256=digest(text.encode()), primary_included=str(included).lower(),
                                            interpretation_status="ANALYST_CODE_OWNER_REVIEW_PENDING",
                                            restricted_pointer=f"restricted://user-research-round1/{fid}/{qid}/{field}"))
    write_json(restricted / "raw-extract.json", raw)
    (restricted / "raw-extract.json").chmod(0o600)
    write_json(restricted / "question-definitions.json", definitions)
    table("USER_TEST_FILE_AUDIT.tsv", list(audit[0]), audit)
    table("USER_TEST_DUPLICATE_CLUSTER.tsv", ["duplicate_cluster_id", "file_sha256", "file_ids", "primary_file_id", "interpretation"],
          [dict(duplicate_cluster_id="D001", file_sha256=next(h for h,ids in clusters.items() if len(ids)>1),
                file_ids="|".join(duplicate[0]), primary_file_id=min(duplicate[0]),
                interpretation="Submitted files do not establish two independent response events; identity is not inferred")])
    table("USER_TEST_RESPONSE_LEDGER.tsv", list(ledger[0]), ledger)
    table("USER_TEST_QUALITATIVE_FEEDBACK.tsv", list(qualitative[0]), qualitative)
    summary, sensitivity = [], []
    for q in range(1,21):
        qid=f"Q{q}"
        prim=Counter(r["coded_answer"] for r in ledger if r["question_id"]==qid and r["primary_included"]=="true")
        all_=Counter(r["coded_answer"] for r in ledger if r["question_id"]==qid)
        for code in ("A","B","C","D","OPEN","MISSING"):
            summary.append(dict(question_id=qid, coded_answer=code, count=prim[code], denominator=10,
                                proportion=f"{prim[code]/10:.6f}", analysis_surface="FILE_DEDUPLICATED"))
            sensitivity.append(dict(question_id=qid, coded_answer=code, primary_count=prim[code], primary_n=10,
                                    raw_count=all_[code], raw_n=11, primary_proportion=f"{prim[code]/10:.6f}",
                                    raw_proportion=f"{all_[code]/11:.6f}"))
    table("USER_TEST_QUESTION_SUMMARY.tsv",list(summary[0]),summary)
    table("USER_TEST_RAW_VS_DEDUP_SENSITIVITY.tsv",list(sensitivity[0]),sensitivity)
    metric_specs = [
        ("COFFEE_WEEKLY_OR_MORE_COUNT",1,"AB","A/B guarantees at least twice weekly; C may include once weekly, so count is a lower bound"),
        ("FLAVOR_DESCRIPTION_EXPOSURE_COUNT",3,"AB","At least sometimes reads descriptions"),
        ("C0_FINDABLE_COUNT",5,"AB","Self-reported findability, not method experience coverage"),
        ("C1_ROUGHLY_SELECTABLE_COUNT",7,"AB","Self-reported selectability, not correct roast classification"),
        ("SEVEN_LEVEL_ACCEPTANCE_COUNT",9,"A","Preference, not validated computational usefulness"),
        ("FIRST_QUESTION_ANSWERABLE_COUNT",10,"AB","Easy or broadly answerable"),
        ("FRUIT_LANGUAGE_DIFFICULTY_COUNT",12,"CD","Only one or two distinctions, or basically cannot distinguish"),
        ("UNLIMITED_MULTISELECT_PREFERENCE_COUNT",15,"C","Preference does not establish information efficiency"),
        ("EXTRA_QUESTION_PREFERENCE_COUNT",19,"B","Prefers additional question as explanation/recovery"),
        ("WILLING_TO_TRY_OR_USE_COUNT",20,"AB","Stated intention; no retention evidence"),
        ("ACTIVE_USE_INTENT_COUNT",20,"A","Stated active-use intention")]
    metrics = {}
    for key,q,codes,note in metric_specs:
        matches=[r for r in ledger if r["question_id"]==f"Q{q}" and r["coded_answer"] in list(codes)]
        metrics[key]=dict(count=sum(r["primary_included"]=="true" for r in matches), denominator=10,
                          raw_count=len(matches), raw_denominator=11, question_id=f"Q{q}", codes=codes, interpretation=note)
    white={r["file_id"] for r in ledger if r["question_id"]=="Q13" and r["coded_answer"]=="A" and r["primary_included"]=="true"}
    white.update(r["file_id"] for r in qualitative if r["qualitative_category"]=="WHITE_FLORAL_NEEDS_EXPLANATION" and r["primary_included"]=="true")
    metrics["WHITE_FLORAL_EXPLANATION_COUNT"]=dict(count=len(white),denominator=10, question_id="Q13",
        interpretation="Union of primary A and explicitly coded open feedback, each file payload counted once")
    aid = restricted / "analytical-review-aid.xlsx"
    comparison=[]
    if not aid.exists():
        raise ValueError("Review aid missing from restricted root")
    ws=load_workbook(aid,data_only=True)["逐题统计"]
    for values in ws.iter_rows(min_row=4,values_only=True):
        if not isinstance(values[0],int): continue
        qid=f"Q{values[0]}"
        for code,col in zip("ABCD",(4,6,8,10)):
            actual=next(r["count"] for r in summary if r["question_id"]==qid and r["coded_answer"]==code)
            comparison.append(dict(question_id=qid,code=code,recomputed_count=actual,review_aid_count=values[col],matches=actual==values[col]))
    if any(not r["matches"] for r in comparison):
        anomaly("REVIEW_AID","ALL","AGGREGATE_DISCREPANCY","ORIGINAL_SOURCE_COUNTS_PREVAIL_SEE_RESTRICTED_RECONCILIATION")
    anomaly("REVIEW_AID","NOTES","UNVERIFIED_REPEATED_DIGIT_CLAIM","NO_SINGLE_DIGIT_NOTE_VALUES_OBSERVED_IN_SUPPLIED_SOURCES; DO_NOT_REPRODUCE_CLAIM")
    write_json(restricted / "review-aid-reconciliation.json",comparison)
    table("USER_TEST_FORMAT_ANOMALY_REGISTER.tsv",list(anomalies[0]),anomalies)
    decision_specs=[
        ("D01","Q5|Q6","Retain eight C0 families; clarify immersion and stovetop examples","C0 self-reported findability; sparse experience outside common methods"),
        ("D02","Q7|Q8|Q9","Retain seven C1 levels and separate unsure state; package-name guidance","No direct reviewed seven-level mappings; C1 remains neutral"),
        ("D03","Q10|Q11|Q12|Q13","Test examples-first and progressive-branch wording between subjects","Small-sample comprehension preferences; neither variant validated"),
        ("D04","Q15","Enable multi-select with a soft 1–2 suggestion; leave unselected options neutral","Preference does not prove efficiency"),
        ("D05","Q17","Keep UNSURE, NONE_OF_THESE and SKIP distinct","Different intended meanings; no single winning label"),
        ("D06","Q16|Q19","Q1 mandatory; conditional Q2–Q4 and optional governed Q5 recovery","Measure burden, separation and acceptance in actual tasks"),
        ("D07","Q18","Up to three headlines with additional main and exploration collapsed","No force-fill; stated preference is not a comparative task result"),
        ("D08","Q20|OVERALL","Add value-proposition paraphrase and repeat-use measurement","Trial interest does not establish retention"),
    ]
    decisions=[dict(decision_id=i,question_ids=q,proposal=p,evidence_limit=l,status="RESEARCH_CANDIDATE_OWNER_REVIEW_PENDING",final_owner_decision="") for i,q,p,l in decision_specs]
    table("USER_TEST_PRODUCT_DECISION.tsv",list(decisions[0]),decisions)
    manifest=dict(artifact_version="user-research-round1",raw_submission_count=len(inv), unique_file_payload_count=len(clusters),
                  duplicate_file_cluster_count=len(duplicate),primary_analysis_n=len(clusters),raw_sensitivity_n=len(inv),
                  open_primary_response_count=sum(r["coded_answer"]=="OPEN" for r in ledger),
                  format_anomaly_count=len(anomalies),single_digit_note_artifact_candidate_count=sum(a["anomaly_type"]=="SINGLE_DIGIT_FORMAT_ARTIFACT_CANDIDATE" for a in anomalies),
                  qualitative_code_count=len(qualitative),numbers_file_included=any(i["format"]=="numbers" for i in inv),
                  study_type="FORMATIVE_SMALL_SAMPLE",population_generalization_allowed=False,
                  user_research_role="FORMATIVE_UX_AND_PRODUCT_POLICY_EVIDENCE",professional_sensory_label_role=False,ml_training_role=False,
                  metrics=metrics,review_aid_sha256=digest(aid.read_bytes()),review_aid_option_count_comparisons=len(comparison),
                  review_aid_option_count_matches=sum(r["matches"] for r in comparison),
                  source_of_truth="Original restricted submissions; native Numbers export used to overcome reader loss",
                  coded_data_scope="Owner-authorized public-safe codes and aggregates; raw files and free text remain restricted",
                  generated_at="2026-09-05T00:00:00Z")
    write_json(OUT/"USER_TEST_MANIFEST.json",manifest)
    checksum="".join(f"{digest(p.read_bytes())}  {p.name}\n" for p in sorted(OUT.iterdir()) if p.is_file() and p.name!="SHA256SUMS")
    (OUT/"SHA256SUMS").write_text(checksum)
    print(json.dumps(dict(raw=len(inv),unique=len(clusters),anomalies=len(anomalies),review_aid_matches=manifest["review_aid_option_count_matches"],metrics={k:v["count"] for k,v in metrics.items()})))


if __name__=="__main__":
    generate()

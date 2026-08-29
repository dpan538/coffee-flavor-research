#!/usr/bin/env python3
"""Acquire and stage professional coffee descriptor candidates for Batch 2.

The network-facing portion writes source artifacts and source-native text only
under an owner-controlled restricted root.  The repository receives a
public-safe sidecar containing stable IDs, hashes, locators, evidence and
rights states, and provisional-normalization hashes.  Re-running with
``--offline`` must reproduce the public outputs byte-for-byte from the cached
restricted artifacts.

This script deliberately does not promote machine decisions to human review,
does not expand frequency counts into judge observations, and does not infer a
descriptor from a score, rank, origin, process, variety, price or award.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import itertools
import json
import re
import sys
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from html.parser import HTMLParser
from pathlib import Path
from typing import Iterable, Iterator

import pdfplumber
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
PUBLIC_DIR = ROOT / "db" / "data" / "professional-descriptor-staging"
BATCH_ID = "professional-descriptor-scaleup-batch2-20260829"
BATCH_DATE = "2026-08-29"
USER_AGENT = "CoffeeFlavorResearch/1.0 public-research-contact"
TARGET_TOTAL = 10_000
MAX_ARTIFACTS = 2_000
MILESTONE_TOTAL = 10_000
HARD_STOP_TOTAL = 20_000
MAX_COE_ARCHIVE_PAGES = 651
OBSERVED_LIVE_ACQUISITION_RUNTIME_SECONDS = 1057.548111
BASELINE_PUBLICATION_OVERLAP_URLS = {
    # Batch 1 already segmented this exact structured detail publication into
    # the 158-row current baseline. Batch 2 retains a false-count audit copy
    # but must not add the publication a second time.
    "https://farmdirectory.cupofexcellence.org/listing/2009-brazil-pulped-naturals-84-96/",
}


STATIC_SOURCES = {
    "zenodo-q-grader-workbook": {
        "url": "https://zenodo.org/records/20840464/files/panelists_scores_EN.xlsx?download=1",
        "filename": "zenodo-20840464-panelists-scores-en.xlsx",
        "family": "family.zenodo_golovinsky_q_grader_dataset",
        "route": "route.zenodo.20840464.panelist-descriptions",
        "schema": "schema.zenodo.panelist-sample-description-matrix.v1",
        "publisher": "Zenodo; Golovinsky et al.",
        "language": "en",
        "rights": "AFFIRMATIVE",
        "rights_basis": "CC_BY_NC_4_0_NONCOMMERCIAL_RESEARCH_ONLY",
        "evidence_tier": "P1",
        "collection_tier": "GOLD",
    },
    "project-origin-current-lots": {
        "url": "https://projectorigin.coffee/intl-home/intl-our-coffee/intl-high-end-lots/",
        "filename": "project-origin-current-lots.html",
        "family": "family.project_origin",
        "route": "route.project-origin.current-lot-table",
        "schema": "schema.professional-green-coffee-table.cupping-notes.v1",
        "publisher": "Project Origin",
        "language": "en",
        "rights": "UNKNOWN",
        "rights_basis": "PUBLIC_DISCOVERY_ONLY_NO_MODEL_PERMISSION_CLAIM",
        "evidence_tier": "P4",
        "collection_tier": "BRONZE",
    },
    "india-fine-cup-2019": {
        "url": "https://hcikl.gov.in/pdf/Winning_Coffees_Brochure.pdf",
        "filename": "india-fine-cup-award-2019.pdf",
        "family": "family.coffee_board_of_india_fine_cup",
        "route": "route.india-fine-cup.2019-winning-coffee-brochure",
        "schema": "schema.official-auction-brochure.one-lot-per-page.v1",
        "publisher": "Coffee Board of India / HCIKL",
        "language": "en",
        "rights": "UNKNOWN",
        "rights_basis": "PUBLIC_DISCOVERY_ONLY_NO_MODEL_PERMISSION_CLAIM",
        "evidence_tier": "UNRESOLVED",
        "collection_tier": "SILVER",
    },
    "isla-hawaii-2021": {
        "url": "https://www.islacoffees.com/wp-content/uploads/2021/03/ISLA.ACE_.COFFEE.AUCTION.2021.pdf",
        "filename": "isla-hawaii-ace-auction-2021.pdf",
        "family": "family.hawaii_agricultural_foundation_isla_ace",
        "route": "route.isla-hawaii.2021-international-jury-auction-booklet",
        "schema": "schema.official-auction-booklet.explicit-international-jury.v1",
        "publisher": "Hawaii Agricultural Foundation / Isla Custom Coffees / ACE",
        "language": "en",
        "rights": "PENDING",
        "rights_basis": "PUBLIC_DISCOVERY_PENDING_REUSE_DECISION",
        "evidence_tier": "P2",
        "collection_tier": "GOLD",
    },
    "fta-green-offering": {
        "url": "https://www.ftacoffee.com.au/_files/ugd/60542b_0acd1804921e408bb1dc18a107316641.pdf?index=true",
        "filename": "fta-green-coffee-offering.pdf",
        "family": "family.forward_specialty_green_coffee",
        "route": "route.fta-green-coffee.offering-list",
        "schema": "schema.professional-green-coffee-pdf.cupping-notes.v1",
        "publisher": "Forward Specialty Green Coffee / FTA Coffee",
        "language": "en",
        "rights": "UNKNOWN",
        "rights_basis": "PUBLIC_DISCOVERY_ONLY_NO_MODEL_PERMISSION_CLAIM",
        "evidence_tier": "P4",
        "collection_tier": "BRONZE",
    },
    "melbourne-coffee-merchants-2024": {
        "url": "https://melbournecoffeemerchants.com.au/wp-content/uploads/2024/11/MCM-Web-List-19.11.24.pdf",
        "filename": "melbourne-coffee-merchants-2024-11-19.pdf",
        "family": "family.melbourne_coffee_merchants",
        "route": "route.mcm.2024-11-19-web-list",
        "schema": "schema.professional-green-coffee-pdf.cupping-notes.v1",
        "publisher": "Melbourne Coffee Merchants",
        "language": "en",
        "rights": "UNKNOWN",
        "rights_basis": "PUBLIC_DISCOVERY_ONLY_NO_MODEL_PERMISSION_CLAIM",
        "evidence_tier": "P4",
        "collection_tier": "BRONZE",
    },
    "sheba-haraz-2021": {
        "url": "https://shebacoffee.com/blogs/auctions-and-listings/haraz-auction-2021",
        "filename": "sheba-haraz-auction-2021.html",
        "family": "family.sheba_coffee_yemen_auction",
        "route": "route.sheba.haraz-auction-2021",
        "schema": "schema.professional-auction-html.cupping-notes.v1",
        "publisher": "Sheba Coffee",
        "language": "en",
        "rights": "UNKNOWN",
        "rights_basis": "PUBLIC_DISCOVERY_ONLY_NO_MODEL_PERMISSION_CLAIM",
        "evidence_tier": "P4",
        "collection_tier": "BRONZE",
    },
}


COE = {
    "family": "family.ace_cup_of_excellence",
    "publisher": "Alliance for Coffee Excellence / Cup of Excellence",
    "rights": "UNKNOWN",
    "rights_basis": "PUBLIC_DISCOVERY_ONLY_NO_MODEL_PERMISSION_CLAIM",
    "language": "en",
}
COE_SEARCH_QUERIES = (
    "Aroma",
    "Flavor",
    "Acidity",
    "Mouthfeel",
    "Overall",
    "Top Jury Descriptions",
)


HISTORICAL_PROBES = (
    {
        "source_family": "family.best_of_panama_scap",
        "publisher": "Specialty Coffee Association of Panama",
        "source_route": "route.best-of-panama.official-result-tables",
        "route_schema": "OFFICIAL_RESULT_TABLE",
        "language": "en",
        "artifacts_inspected": 14,
        "descriptor_bearing_artifacts": 0,
        "descriptor_bearing_records": 0,
        "raw_assertions": 0,
        "deinflated_assertions": 0,
        "strict_assertions": 0,
        "broad_assertions": 0,
        "gold_assertions": 0,
        "silver_assertions": 0,
        "bronze_assertions": 0,
        "duplicate_losses": 0,
        "non_descriptor_losses": 408,
        "rights_note": "PENDING",
        "disposition": "ZERO_YIELD_ROUTE_STOPPED_RESULT_METADATA_ONLY",
        "continuation_cursor": "official-scap-archive-2003",
        "basis": "ROUND3L_PUBLIC_CHECKPOINT_REUSED_NO_RESCAN",
    },
    {
        "source_family": "family.world_coffee_events",
        "publisher": "World Coffee Events / national Competition Bodies",
        "source_route": "route.wcc.public-results-and-blank-scoresheets",
        "route_schema": "RESULT_OR_BLANK_FORM",
        "language": "multi",
        "artifacts_inspected": 29,
        "descriptor_bearing_artifacts": 0,
        "descriptor_bearing_records": 0,
        "raw_assertions": 0,
        "deinflated_assertions": 0,
        "strict_assertions": 0,
        "broad_assertions": 0,
        "gold_assertions": 0,
        "silver_assertions": 0,
        "bronze_assertions": 0,
        "duplicate_losses": 7,
        "non_descriptor_losses": 466,
        "rights_note": "UNKNOWN",
        "disposition": "ZERO_YIELD_PUBLIC_RESULT_AND_BLANK_FORM_ROUTE",
        "continuation_cursor": "source:wcc_cb_recdwegkkkkomky5z",
        "basis": "ROUND3L_PUBLIC_CHECKPOINT_REUSED_NO_RESCAN",
    },
    {
        "source_family": "family.golden_bean",
        "publisher": "Golden Bean",
        "source_route": "route.golden-bean.public-winner-results",
        "route_schema": "AWARD_RESULTS",
        "language": "en",
        "artifacts_inspected": 57,
        "descriptor_bearing_artifacts": 0,
        "descriptor_bearing_records": 0,
        "raw_assertions": 0,
        "deinflated_assertions": 0,
        "strict_assertions": 0,
        "broad_assertions": 0,
        "gold_assertions": 0,
        "silver_assertions": 0,
        "bronze_assertions": 0,
        "duplicate_losses": 337,
        "non_descriptor_losses": 2419,
        "rights_note": "UNKNOWN",
        "disposition": "ZERO_YIELD_ROUTE_STOPPED_AWARDS_ONLY",
        "continuation_cursor": "EXHAUSTED_PUBLIC_RESULT_SCHEMA",
        "basis": "ROUND3L_PUBLIC_CHECKPOINT_REUSED_NO_RESCAN",
    },
    {
        "source_family": "family.coffee_quality_institute",
        "publisher": "Coffee Quality Institute",
        "source_route": "route.cqi.public-sample-grade-detail",
        "route_schema": "PROFESSIONAL_SAMPLE_DETAIL",
        "language": "en",
        "artifacts_inspected": 1,
        "descriptor_bearing_artifacts": 0,
        "descriptor_bearing_records": 0,
        "raw_assertions": 0,
        "deinflated_assertions": 0,
        "strict_assertions": 0,
        "broad_assertions": 0,
        "gold_assertions": 0,
        "silver_assertions": 0,
        "bronze_assertions": 0,
        "duplicate_losses": 0,
        "non_descriptor_losses": 0,
        "rights_note": "UNKNOWN",
        "disposition": "ZERO_YIELD_VISIBLE_FILLED_VALUES_ABSENT",
        "continuation_cursor": "WAIT_FOR_ROUTE_SCHEMA_CHANGE",
        "basis": "BATCH1_ROLLING_RECEIPT_REUSED_NO_RESCAN",
    },
    {
        "source_family": "family.iiac_international_coffee_tasting",
        "publisher": "International Institute of Coffee Tasters",
        "source_route": "route.iiac.public-competition-methodology",
        "route_schema": "METHODOLOGY_AND_MEDAL_LIST",
        "language": "en|it",
        "artifacts_inspected": 3,
        "descriptor_bearing_artifacts": 0,
        "descriptor_bearing_records": 0,
        "raw_assertions": 0,
        "deinflated_assertions": 0,
        "strict_assertions": 0,
        "broad_assertions": 0,
        "gold_assertions": 0,
        "silver_assertions": 0,
        "bronze_assertions": 0,
        "duplicate_losses": 0,
        "non_descriptor_losses": 0,
        "rights_note": "UNKNOWN",
        "disposition": "REFERENCE_ONLY_PUBLIC_PROFILE_VALUES_NOT_EXPOSED",
        "continuation_cursor": "PARTNERSHIP_REGISTER_ONLY_NO_REQUEST_SENT",
        "basis": "DESCRIPTOR_CENSUS_PDF_AND_PUBLIC_ROUTE_PROBE",
    },
    {
        "source_family": "family.afca_taste_of_harvest",
        "publisher": "African Fine Coffees Association",
        "source_route": "route.afca.taste-of-harvest-public-results",
        "route_schema": "RESULT_AND_PROGRAM_SUMMARY",
        "language": "en",
        "artifacts_inspected": 3,
        "descriptor_bearing_artifacts": 0,
        "descriptor_bearing_records": 0,
        "raw_assertions": 0,
        "deinflated_assertions": 0,
        "strict_assertions": 0,
        "broad_assertions": 0,
        "gold_assertions": 0,
        "silver_assertions": 0,
        "bronze_assertions": 0,
        "duplicate_losses": 0,
        "non_descriptor_losses": 0,
        "rights_note": "UNKNOWN",
        "disposition": "ZERO_YIELD_GENERAL_PROFILE_NO_LOT_LEVEL_OBSERVATIONS",
        "continuation_cursor": "NEXT_OFFICIAL_AUCTION_LOT_ROUTE",
        "basis": "PUBLIC_ROUTE_PROBE_2026-08-29",
    },
    {
        "source_family": "family.coffee_review",
        "publisher": "Coffee Review",
        "source_route": "route.coffee-review.professional-review-pages",
        "route_schema": "PROFESSIONAL_BLIND_ASSESSMENT",
        "language": "en",
        "artifacts_inspected": 0,
        "descriptor_bearing_artifacts": 0,
        "descriptor_bearing_records": 0,
        "raw_assertions": 0,
        "deinflated_assertions": 0,
        "strict_assertions": 0,
        "broad_assertions": 0,
        "gold_assertions": 0,
        "silver_assertions": 0,
        "bronze_assertions": 0,
        "duplicate_losses": 0,
        "non_descriptor_losses": 0,
        "rights_note": "UNKNOWN",
        "disposition": "BLOCKED_ROBOTS_DO_NOT_ACQUIRE",
        "continuation_cursor": "NONE_UNLESS_ACCESS_POLICY_CHANGES",
        "basis": "SEARCH_DISCOVERY_ONLY_ROBOTS_BLOCK_OBSERVED",
    },
)


BROAD_EXACT = {
    "above average",
    "above middle",
    "acidic",
    "acidity",
    "aftertaste",
    "average",
    "balanced",
    "below average",
    "below middle",
    "bitter",
    "bitterness",
    "body",
    "bright",
    "caustic",
    "clean",
    "cleanliness",
    "complex",
    "consistent",
    "creamy",
    "crisp",
    "delicate",
    "dense",
    "dries",
    "dry",
    "enveloping",
    "fruity",
    "full",
    "high",
    "intense",
    "juicy",
    "light",
    "lingering",
    "long",
    "long lasting",
    "low",
    "medium",
    "middle",
    "missing",
    "pleasant",
    "refined",
    "rich",
    "rough",
    "round",
    "short",
    "silky",
    "smooth",
    "sparkling",
    "sparkly",
    "sweet",
    "sweetness",
    "syrupy",
    "tart",
    "transparent",
    "unique",
    "velvety",
    "vibrant",
    "winey",
}
BROAD_SUFFIXES = (
    " acidity",
    " aftertaste",
    " body",
    " finish",
    " mouthfeel",
    " sweetness",
    " bitterness",
    " texture",
)
NON_DESCRIPTOR = {"", "-", "missing", "n/a", "na", "none", "not detected"}
ALIASES = {
    "berries": "berry",
    "blackcurrant": "black currant",
    "chocolaty": "chocolate",
    "chocolatey": "chocolate",
    "flowers": "flower",
    "fruits": "fruit",
    "mallic acidity": "malic acidity",
    "nuts": "nut",
    "raisins": "raisin",
    "spices": "spice",
}


@dataclass
class Atom:
    source_family: str
    publisher: str
    source_route: str
    route_schema: str
    source_url: str
    artifact_sha256: str
    source_locator: str
    effective_record_id: str
    coffee_identity_id: str
    edition_or_release: str
    edition_year: str
    source_language: str
    preparation_service: str
    roast_evidence: str
    source_field_label: str
    raw_field_text: str
    atomic_text: str
    evidence_tier: str
    collection_tier: str
    provenance_state: str
    rights_state: str
    rights_basis: str
    publication_layer: str
    judge_observation_id: str = ""
    descriptor_class: str = ""
    provisional_form: str = ""
    provisional_form_id: str = ""
    mapping_status: str = ""
    mapping_method: str = ""
    mapping_confidence: str = ""
    mapping_basis: str = ""
    review_requirement: str = ""
    counts_as_assertion: bool = True
    counts_as_record_unique_descriptor: bool = True
    descriptor_assertion_id: str = ""


@dataclass
class RouteMetric:
    source_family: str
    publisher: str
    source_route: str
    route_schema: str
    language: str
    rights_note: str
    artifacts_inspected: int = 0
    descriptor_bearing_artifacts: int = 0
    descriptor_bearing_records: int = 0
    raw_assertions: int = 0
    deinflated_assertions: int = 0
    strict_assertions: int = 0
    broad_assertions: int = 0
    gold_assertions: int = 0
    silver_assertions: int = 0
    bronze_assertions: int = 0
    duplicate_losses: int = 0
    non_descriptor_losses: int = 0
    analyst_equivalent_minutes: str = "NA_AUTOMATED_ROUTE_NO_MANUAL_TIMING"
    automated_runtime_seconds: str = ""
    disposition: str = "DISCOVERED"
    continuation_cursor: str = ""
    basis: str = "BATCH2_DIRECT_ROUTE"
    artifact_ids: set[str] = field(default_factory=set)
    record_ids: set[str] = field(default_factory=set)


class TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self._row: list[str] | None = None
        self._cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "tr":
            self._row = []
        elif tag in {"th", "td"} and self._row is not None:
            self._cell = []
        elif tag == "br" and self._cell is not None:
            self._cell.append(" ")

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            self._cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag in {"th", "td"} and self._row is not None and self._cell is not None:
            self._row.append(space("".join(self._cell)))
            self._cell = None
        elif tag == "tr" and self._row is not None:
            if self._row:
                self.rows.append(self._row)
            self._row = None


class CoeParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.fields: dict[str, str] = {}
        self._kind = ""
        self._buf: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        classes = set(dict(attrs).get("class", "").split())
        if tag == "div" and "item-attr" in classes:
            self._kind, self._buf = "label", []
        elif tag == "div" and "item-property" in classes:
            self._kind, self._buf = "value", []

    def handle_data(self, data: str) -> None:
        if self._kind:
            self._buf.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag != "div" or not self._kind:
            return
        value = space("".join(self._buf))
        if self._kind == "label":
            self._last_label = value
        elif getattr(self, "_last_label", ""):
            self.fields[self._last_label] = value
        self._kind, self._buf = "", []


def space(value: str) -> str:
    return " ".join(value.replace("\xa0", " ").split())


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_file(path: Path) -> str:
    return sha256_bytes(path.read_bytes())


def stable_id(prefix: str, *parts: object) -> str:
    material = "\x1f".join(str(part) for part in parts)
    return f"{prefix}:{hashlib.sha256(material.encode('utf-8')).hexdigest()[:24]}"


def normalize(value: str) -> tuple[str, str, str, str, str]:
    native = space(unicodedata.normalize("NFKC", value).strip(" .,:;\t\n"))
    provisional = native.casefold().replace("’", "'")
    provisional = re.sub(r"\s*([/&])\s*", r" \1 ", provisional)
    provisional = space(provisional)
    status = "AUTO_ORTHOGRAPHIC"
    method = "NFKC_CASE_SPACE_PUNCTUATION"
    basis = "CASE_SPACING_AND_PUNCTUATION_NORMALIZATION"
    confidence = "0.990000"
    if provisional in ALIASES:
        provisional = ALIASES[provisional]
        status = "AUTO_APPROVED_ALIAS"
        method = "DOCUMENTED_ORTHOGRAPHIC_ALIAS"
        basis = "BATCH2_APPROVED_SPELLING_OR_PLURAL_ALIAS"
        confidence = "0.990000"
    if re.match(r"^[a-z]-\s*", provisional) or " and " in provisional:
        status = "MACHINE_SUGGESTED_REVIEW_REQUIRED"
        method = "CONSERVATIVE_COMPOUND_PRESERVATION"
        basis = "AMBIGUOUS_PREFIX_OR_COMPOUND_RETAINED_INTACT"
        confidence = "0.650000"
    return native, provisional, status, method, confidence + "|" + basis


def descriptor_class(value: str) -> str:
    folded = value.casefold()
    if folded in NON_DESCRIPTOR:
        return "NON_DESCRIPTOR"
    if folded in BROAD_EXACT or folded.endswith(BROAD_SUFFIXES):
        return "BROAD_SENSORY"
    if any(token in folded for token in ("balanced", "clean", "finish", "mouthfeel")):
        return "BROAD_SENSORY"
    return "STRICT_FLAVOR"


def split_atomic(value: str) -> tuple[str, ...]:
    result: list[str] = []
    for part in re.split(r"[,;]", value):
        item = space(part.strip(" .,:;\t\n"))
        if item:
            result.append(item)
    return tuple(result)


def make_atoms(
    *,
    source: dict[str, str],
    artifact_sha256: str,
    source_url: str,
    source_locator: str,
    effective_record_id: str,
    coffee_identity_id: str,
    edition_or_release: str,
    edition_year: str,
    preparation_service: str,
    roast_evidence: str,
    source_field_label: str,
    raw_field_text: str,
    publication_layer: str,
    provenance_state: str,
    judge_observation_id: str = "",
) -> list[Atom]:
    atoms: list[Atom] = []
    for occurrence, raw_atom in enumerate(split_atomic(raw_field_text)):
        native, provisional, status, method, confidence_basis = normalize(raw_atom)
        confidence, basis = confidence_basis.split("|", 1)
        klass = descriptor_class(provisional)
        atom = Atom(
            source_family=source["family"],
            publisher=source["publisher"],
            source_route=source["route"],
            route_schema=source["schema"],
            source_url=source_url,
            artifact_sha256=artifact_sha256,
            source_locator=source_locator,
            effective_record_id=effective_record_id,
            coffee_identity_id=coffee_identity_id,
            edition_or_release=edition_or_release,
            edition_year=edition_year,
            source_language=source["language"],
            preparation_service=preparation_service,
            roast_evidence=roast_evidence,
            source_field_label=source_field_label,
            raw_field_text=raw_field_text,
            atomic_text=native,
            evidence_tier=source["evidence_tier"],
            collection_tier=source["collection_tier"],
            provenance_state=provenance_state,
            rights_state=source["rights"],
            rights_basis=source["rights_basis"],
            publication_layer=publication_layer,
            judge_observation_id=judge_observation_id,
            descriptor_class=klass,
            provisional_form=provisional,
            provisional_form_id=stable_id("provisional-form", provisional),
            mapping_status=status,
            mapping_method=method,
            mapping_confidence=confidence,
            mapping_basis=basis,
            review_requirement=(
                "SENSORY_EXPERT_ADJUDICATION_REQUIRED"
                if status == "MACHINE_SUGGESTED_REVIEW_REQUIRED"
                else "AUTO_PROVISIONALLY_MAPPED"
            ),
        )
        atom.descriptor_assertion_id = stable_id(
            "assertion-b2",
            atom.source_route,
            atom.artifact_sha256,
            atom.effective_record_id,
            atom.publication_layer,
            atom.source_locator,
            atom.judge_observation_id,
            atom.provisional_form,
            occurrence,
        )
        atoms.append(atom)
    return atoms


def fetch(url: str, path: Path, *, offline: bool, delay: float = 0.0) -> bool:
    if path.is_file() and path.stat().st_size:
        return False
    if offline:
        raise RuntimeError(f"offline cache is missing: {path}")
    if delay:
        time.sleep(delay)
    path.parent.mkdir(parents=True, exist_ok=True)
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(request, timeout=45) as response:
            payload = response.read()
    except (urllib.error.URLError, TimeoutError) as exc:
        raise RuntimeError(f"failed to fetch {url}: {exc}") from exc
    path.write_bytes(payload)
    return True


def apply_deinflation(atoms: list[Atom]) -> tuple[int, int]:
    observation_seen: set[tuple[str, ...]] = set()
    record_seen: set[tuple[str, str]] = set()
    assertion_losses = 0
    record_losses = 0
    for atom in atoms:
        observation_key = (
            atom.effective_record_id,
            atom.artifact_sha256,
            atom.publication_layer,
            atom.source_locator,
            atom.judge_observation_id,
            atom.provisional_form_id,
        )
        if atom.descriptor_class == "NON_DESCRIPTOR" or observation_key in observation_seen:
            atom.counts_as_assertion = False
            assertion_losses += 1
        else:
            observation_seen.add(observation_key)
        record_key = (atom.effective_record_id, atom.provisional_form_id)
        if not atom.counts_as_assertion or record_key in record_seen:
            atom.counts_as_record_unique_descriptor = False
            if atom.counts_as_assertion:
                record_losses += 1
        else:
            record_seen.add(record_key)
    return assertion_losses, record_losses


def suppress_baseline_publication_overlap(atoms: list[Atom], source_url: str) -> int:
    if source_url not in BASELINE_PUBLICATION_OVERLAP_URLS:
        return 0
    losses = 0
    for atom in atoms:
        if atom.counts_as_assertion:
            losses += 1
        atom.counts_as_assertion = False
        atom.counts_as_record_unique_descriptor = False
    return losses


def parse_zenodo(source: dict[str, str], path: Path) -> list[list[Atom]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["All Panelists"]
    description_columns = {
        4: "Aroma description",
        6: "Bouquet description",
        8: "Aftertaste description",
        10: "Acidity description",
        13: "Sweetness description",
        16: "Bitterness description",
        19: "Body description",
    }
    result: list[list[Atom]] = []
    artifact_hash = sha256_file(path)
    for row_number, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), 3):
        sample_number, panelist, sample_name, grind_roast = row[:4]
        if sample_number is None or not panelist or not sample_name:
            continue
        sample_key = str(sample_number)
        effective = stable_id("effective-b2", source["route"], sample_key, "SCA_PROTOCOL")
        coffee = stable_id("coffee-b2", source["route"], sample_key)
        judge = stable_id("panelist-b2", panelist)
        roast = space(str(grind_roast or "")) or "UNREPORTED"
        record_atoms: list[Atom] = []
        for index, label in description_columns.items():
            value = row[index]
            if not isinstance(value, str) or not value.strip():
                continue
            record_atoms.extend(
                make_atoms(
                    source=source,
                    artifact_sha256=artifact_hash,
                    source_url=source["url"],
                    source_locator=f"sheet:All Panelists#row={row_number};column={index + 1}",
                    effective_record_id=effective,
                    coffee_identity_id=coffee,
                    edition_or_release="Zenodo record 20840464 v1.1",
                    edition_year="2026",
                    preparation_service="CUPPING",
                    roast_evidence=roast,
                    source_field_label=label,
                    raw_field_text=value,
                    publication_layer="JUDGE_LEVEL_OBSERVATION",
                    provenance_state="SOURCE_AUDITED_EXPLICIT_Q_GRADER_OBSERVATION",
                    judge_observation_id=judge,
                )
            )
        if record_atoms:
            result.append(record_atoms)
    return result


def parse_project_origin(source: dict[str, str], path: Path) -> list[list[Atom]]:
    parser = TableParser()
    parser.feed(path.read_text(encoding="utf-8", errors="replace"))
    artifact_hash = sha256_file(path)
    result: list[list[Atom]] = []
    header: list[str] | None = None
    for row_index, row in enumerate(parser.rows, 1):
        lowered = [item.casefold() for item in row]
        if "cupping notes" in lowered:
            header = lowered
            continue
        if not header or len(row) != len(header):
            continue
        notes_index = header.index("cupping notes")
        notes = row[notes_index]
        if not notes:
            continue
        coffee_index = header.index("coffee") if "coffee" in header else 1
        year_index = header.index("year") if "year" in header else -1
        record_name = row[coffee_index] or f"row-{row_index}"
        year = row[year_index] if year_index >= 0 else "2025-2026"
        effective = stable_id("effective-b2", source["route"], record_name, year)
        atoms = make_atoms(
            source=source,
            artifact_sha256=artifact_hash,
            source_url=source["url"],
            source_locator=f"html:table-row={row_index};field=Cupping Notes",
            effective_record_id=effective,
            coffee_identity_id=stable_id("coffee-b2", source["route"], record_name),
            edition_or_release="Current exceptional and competition lots",
            edition_year=(re.search(r"20\d{2}", year) or [""])[0],
            preparation_service="CUPPING",
            roast_evidence="UNREPORTED",
            source_field_label="Cupping Notes",
            raw_field_text=notes,
            publication_layer="PROFESSIONAL_COMMERCIAL_SENSORY_FIELD",
            provenance_state="PROFESSIONAL_PUBLISHER_FIELD_AUTHOR_UNRESOLVED",
        )
        if atoms:
            result.append(atoms)
    return result


def parse_india_pdf(source: dict[str, str], path: Path) -> list[list[Atom]]:
    result: list[list[Atom]] = []
    artifact_hash = sha256_file(path)
    with pdfplumber.open(path) as pdf:
        for page_number, page in enumerate(pdf.pages, 1):
            text = page.extract_text(layout=True) or ""
            rank = re.search(r"Rank\s+(\d+[A-Za-z]?)", text, re.I)
            score = re.search(r"Score\s+\d", text, re.I)
            if not rank or not score:
                continue
            normalized = space(text)
            match = re.search(
                r"Weight\s*\(Kg\)\s*(.*?)(?:For auction|Third Party Professional Audit Agency)",
                normalized,
                re.I,
            )
            if not match:
                continue
            notes = space(match.group(1))
            notes = re.sub(r"^\d+(?:\.\d+)?\s*", "", notes)
            if len(notes) < 3:
                continue
            record = f"rank-{rank.group(1).casefold()}-page-{page_number}"
            effective = stable_id("effective-b2", source["route"], record)
            atoms = make_atoms(
                source=source,
                artifact_sha256=artifact_hash,
                source_url=source["url"],
                source_locator=f"pdf:page={page_number};field=tasting-note-box",
                effective_record_id=effective,
                coffee_identity_id=stable_id("coffee-b2", source["route"], record),
                edition_or_release="Fine Cup Award 2019",
                edition_year="2019",
                preparation_service="CUPPING",
                roast_evidence="UNREPORTED",
                source_field_label="Tasting note",
                raw_field_text=notes,
                publication_layer="OFFICIAL_AUCTION_LOT_SENSORY_FIELD",
                provenance_state="OFFICIAL_FIELD_ORIGIN_UNRESOLVED",
            )
            if atoms:
                result.append(atoms)
    return result


def parse_isla_pdf(source: dict[str, str], path: Path) -> list[list[Atom]]:
    result: list[list[Atom]] = []
    artifact_hash = sha256_file(path)
    with pdfplumber.open(path) as pdf:
        for page_number, page in enumerate(pdf.pages, 1):
            text = page.extract_text(layout=True) or ""
            rank = re.search(r"RANK\s*:\s*NO\.\s*([^\s]+)", text, re.I)
            match = re.search(r"Tasting Notes\s*:\s*(.*?)Estimated Lot Size", text, re.I | re.S)
            if not rank or not match:
                continue
            notes = space(match.group(1))
            record = f"rank-{rank.group(1).casefold()}-page-{page_number}"
            effective = stable_id("effective-b2", source["route"], record)
            atoms = make_atoms(
                source=source,
                artifact_sha256=artifact_hash,
                source_url=source["url"],
                source_locator=f"pdf:page={page_number};field=Tasting Notes",
                effective_record_id=effective,
                coffee_identity_id=stable_id("coffee-b2", source["route"], record),
                edition_or_release="Hawaii ACE Coffee Auction 2021",
                edition_year="2021",
                preparation_service="CUPPING",
                roast_evidence="UNREPORTED",
                source_field_label="Tasting Notes",
                raw_field_text=notes,
                publication_layer="PRIMARY_JURY_DESCRIPTION",
                provenance_state="SOURCE_AUDITED_EXPLICIT_INTERNATIONAL_JURY_FIELD",
            )
            if atoms:
                result.append(atoms)
    return result


def parse_generic_cupping_pdf(source: dict[str, str], path: Path) -> list[list[Atom]]:
    result: list[list[Atom]] = []
    artifact_hash = sha256_file(path)
    with pdfplumber.open(path) as pdf:
        for page_number, page in enumerate(pdf.pages, 1):
            text = page.extract_text(layout=True) or ""
            for match_index, match in enumerate(
                re.finditer(
                    r"(?:CUPPING NOTES|Cupping Notes)\s*[:|]?\s*(.*?)(?=(?:QTY|Qty|PACKAGING|Packaging|Availability|\n\s*\n|$))",
                    text,
                    re.S,
                ),
                1,
            ):
                notes = space(match.group(1))
                if len(notes) < 3 or notes.casefold() == "none":
                    continue
                record = f"page-{page_number}-notes-{match_index}"
                effective = stable_id("effective-b2", source["route"], record)
                atoms = make_atoms(
                    source=source,
                    artifact_sha256=artifact_hash,
                    source_url=source["url"],
                    source_locator=f"pdf:page={page_number};field=Cupping Notes;index={match_index}",
                    effective_record_id=effective,
                    coffee_identity_id=stable_id("coffee-b2", source["route"], record),
                    edition_or_release=path.stem,
                    edition_year=(re.search(r"20\d{2}", path.stem) or [""])[0],
                    preparation_service="CUPPING",
                    roast_evidence="UNREPORTED",
                    source_field_label="Cupping Notes",
                    raw_field_text=notes,
                    publication_layer="PROFESSIONAL_COMMERCIAL_SENSORY_FIELD",
                    provenance_state="PROFESSIONAL_PUBLISHER_FIELD_AUTHOR_UNRESOLVED",
                )
                if atoms:
                    result.append(atoms)
    return result


def html_text(path: Path) -> str:
    value = path.read_text(encoding="utf-8", errors="replace")
    value = re.sub(r"<script\b.*?</script>", " ", value, flags=re.I | re.S)
    value = re.sub(r"<style\b.*?</style>", " ", value, flags=re.I | re.S)
    value = re.sub(r"</(?:p|div|h[1-6]|li|tr)>", "\n", value, flags=re.I)
    value = re.sub(r"<br\s*/?>", "\n", value, flags=re.I)
    value = re.sub(r"<[^>]+>", " ", value)
    return html.unescape(value)


def parse_sheba(source: dict[str, str], path: Path) -> list[list[Atom]]:
    text = html_text(path)
    pattern = re.compile(
        r"Auction Lot\s*:\s*([^\n]+).*?Cupping Notes\s*:\s*(.*?)(?=Auction Lot\s*:|$)",
        re.I | re.S,
    )
    result: list[list[Atom]] = []
    artifact_hash = sha256_file(path)
    for index, match in enumerate(pattern.finditer(text), 1):
        lot = space(match.group(1))
        notes = space(match.group(2).split("Estimated", 1)[0])
        if not notes:
            continue
        effective = stable_id("effective-b2", source["route"], lot)
        atoms = make_atoms(
            source=source,
            artifact_sha256=artifact_hash,
            source_url=source["url"],
            source_locator=f"html:auction-lot={index};field=Cupping Notes",
            effective_record_id=effective,
            coffee_identity_id=stable_id("coffee-b2", source["route"], lot),
            edition_or_release="Haraz Auction 2021",
            edition_year="2021",
            preparation_service="CUPPING",
            roast_evidence="UNREPORTED",
            source_field_label="Cupping Notes",
            raw_field_text=notes,
            publication_layer="PROFESSIONAL_AUCTION_LOT_SENSORY_FIELD",
            provenance_state="PROFESSIONAL_PUBLISHER_FIELD_AUTHOR_UNRESOLVED",
        )
        if atoms:
            result.append(atoms)
    return result


def coe_search_url(query: str, page: int) -> str:
    encoded = urllib.parse.quote_plus(query)
    if page == 1:
        return f"https://farmdirectory.cupofexcellence.org/?s={encoded}"
    return f"https://farmdirectory.cupofexcellence.org/page/{page}/?s={encoded}"


def coe_search_links(payload: str) -> tuple[set[str], int]:
    links = set(
        re.findall(
            r'href="(https://farmdirectory\.cupofexcellence\.org/listing/[^"#?]+/)"',
            payload,
            re.I,
        )
    )
    pages = [int(value) for value in re.findall(r"Page\s+(\d+)", payload)]
    return links, max(pages or [1])


def coe_archive_url(page: int) -> str:
    if page == 1:
        return "https://farmdirectory.cupofexcellence.org/listings/"
    return f"https://farmdirectory.cupofexcellence.org/listings/page/{page}/"


def coe_archive_links(payload: str) -> set[str]:
    return set(
        re.findall(
            r'href="(https://farmdirectory\.cupofexcellence\.org/listing/[^"#?]+/)"',
            payload,
            re.I,
        )
    )


def strip_tags(value: str) -> str:
    return space(html.unescape(re.sub(r"<[^>]+>", " ", value)))


def split_jury_sections(value: str) -> list[tuple[str, str]]:
    pattern = re.compile(
        r"(Aroma\s*/?\s*Flavor|Flavor\s*/?\s*Aroma|Aroma|Flavor|Acidity|Other\s+features?|Mouthfeel)\s*:\s*",
        re.I,
    )
    matches = list(pattern.finditer(value))
    sections: list[tuple[str, str]] = []
    for index, match in enumerate(matches):
        end = matches[index + 1].start() if index + 1 < len(matches) else len(value)
        sections.append((space(match.group(1)), space(value[match.end() : end])))
    return sections


def parse_coe_detail(path: Path, source_url: str) -> list[Atom]:
    raw = path.read_text(encoding="utf-8", errors="replace")
    parser = CoeParser()
    parser.feed(raw)
    fields = parser.fields
    year = fields.get("Year", "")
    slug = urllib.parse.urlparse(source_url).path.rstrip("/").rsplit("/", 1)[-1]
    effective = stable_id("effective-b2", "route.coe.detail", slug)
    coffee = stable_id("coffee-b2", "family.ace_cup_of_excellence", slug)
    artifact_hash = sha256_file(path)
    top_jury_match = re.search(
        r"Top Jury Descriptions\s*:\s*</strong>\s*(.*?)</td>", raw, re.I | re.S
    )
    score_pair = bool(re.search(r"Score from International Judges\s*", raw, re.I))
    atoms: list[Atom] = []
    if top_jury_match and score_pair:
        source = {
            **COE,
            "route": "route.coe.explicit-top-jury-detail",
            "schema": "schema.coe.explicit-top-jury-description.v2",
            "evidence_tier": "P2",
            "collection_tier": "GOLD",
            "rights": "PENDING",
            "rights_basis": "PUBLIC_DISCOVERY_PENDING_PURPOSE_SPECIFIC_REUSE_DECISION",
        }
        jury_value = strip_tags(top_jury_match.group(1))
        sections = split_jury_sections(jury_value)
        for section_index, (label, value) in enumerate(sections, 1):
            atoms.extend(
                make_atoms(
                    source=source,
                    artifact_sha256=artifact_hash,
                    source_url=source_url,
                    source_locator=f"html:Top Jury Descriptions;section={section_index}",
                    effective_record_id=effective,
                    coffee_identity_id=coffee,
                    edition_or_release=f"Cup of Excellence {year or 'undated'}",
                    edition_year=year,
                    preparation_service="CUPPING",
                    roast_evidence="UNREPORTED",
                    source_field_label=label,
                    raw_field_text=value,
                    publication_layer="PRIMARY_JURY_DESCRIPTION",
                    provenance_state="SOURCE_AUDITED_EXPLICIT_JURY_FIELD",
                )
            )
        return atoms

    source = {
        **COE,
        "route": "route.coe.generic-official-sensory-detail",
        "schema": "schema.coe.structured-official-sensory-fields.v2",
        "evidence_tier": "UNRESOLVED",
        "collection_tier": "SILVER",
    }
    for field_index, label in enumerate(("Overall", "Aroma / Flavor", "Acidity", "Mouthfeel / Other", "Other"), 1):
        value = fields.get(label, "")
        if not value:
            continue
        frequency_coded = bool(re.search(r"\([^)]*\d+[^)]*\)", value))
        atoms.extend(
            make_atoms(
                source=source,
                artifact_sha256=artifact_hash,
                source_url=source_url,
                source_locator=f"html:item-attr={label};index={field_index}",
                effective_record_id=effective,
                coffee_identity_id=coffee,
                edition_or_release=f"Cup of Excellence {year or 'undated'}",
                edition_year=year,
                preparation_service="CUPPING",
                roast_evidence="UNREPORTED",
                source_field_label=label,
                raw_field_text=value,
                publication_layer="GENERIC_ORGANIZER_SENSORY_FIELD",
                provenance_state=(
                    "FREQUENCY_CODED_OFFICIAL_FIELD_ORIGIN_UNRESOLVED"
                    if frequency_coded
                    else "OFFICIAL_FIELD_ORIGIN_UNRESOLVED"
                ),
            )
        )
    return atoms


def safe_rows(atoms: Iterable[Atom]) -> Iterator[dict[str, object]]:
    for atom in atoms:
        yield {
            "descriptor_assertion_id": atom.descriptor_assertion_id,
            "source_family_id": atom.source_family,
            "publisher": atom.publisher,
            "source_route_id": atom.source_route,
            "route_schema": atom.route_schema,
            "source_url": atom.source_url,
            "source_artifact_sha256": atom.artifact_sha256,
            "source_locator": atom.source_locator,
            "effective_record_id": atom.effective_record_id,
            "coffee_identity_id": atom.coffee_identity_id,
            "edition_or_release": atom.edition_or_release,
            "edition_year": atom.edition_year,
            "source_language": atom.source_language,
            "preparation_service": atom.preparation_service,
            "roast_evidence_sha256_or_state": (
                atom.roast_evidence
                if atom.roast_evidence in {"UNREPORTED", "SOURCE_UNKNOWN"}
                else "hash:sha256:" + sha256_bytes(atom.roast_evidence.encode("utf-8"))
            ),
            "source_field_label": atom.source_field_label,
            "raw_field_text_sha256": sha256_bytes(atom.raw_field_text.encode("utf-8")),
            "atomic_source_text_sha256": sha256_bytes(atom.atomic_text.encode("utf-8")),
            "source_native_form_sha256": sha256_bytes(atom.atomic_text.encode("utf-8")),
            "descriptor_class": atom.descriptor_class,
            "evidence_tier": atom.evidence_tier,
            "collection_tier": atom.collection_tier,
            "provenance_state": atom.provenance_state,
            "rights_state": atom.rights_state,
            "rights_basis": atom.rights_basis,
            "publication_layer": atom.publication_layer,
            "judge_observation_id_sha256": (
                sha256_bytes(atom.judge_observation_id.encode("utf-8"))
                if atom.judge_observation_id
                else ""
            ),
            "provisional_normalized_form_id": atom.provisional_form_id,
            "provisional_normalized_form_sha256": sha256_bytes(
                atom.provisional_form.encode("utf-8")
            ),
            "mapping_status": atom.mapping_status,
            "mapping_method": atom.mapping_method,
            "mapping_confidence": atom.mapping_confidence,
            "mapping_basis": atom.mapping_basis,
            "review_requirement": atom.review_requirement,
            "human_reviewed": False,
            "model_eligible": False,
            "counts_as_assertion": atom.counts_as_assertion,
            "counts_as_record_unique_descriptor": atom.counts_as_record_unique_descriptor,
            "source_text_storage_state": "OWNER_CONTROLLED_RESTRICTED_HASH_ONLY_PUBLIC",
        }


def restricted_rows(atoms: Iterable[Atom]) -> Iterator[dict[str, object]]:
    for atom in atoms:
        row = next(safe_rows((atom,)))
        row.update(
            raw_field_text=atom.raw_field_text,
            atomic_source_text=atom.atomic_text,
            source_native_form=atom.atomic_text,
            provisional_normalized_form=atom.provisional_form,
            judge_observation_id=atom.judge_observation_id,
            roast_evidence=atom.roast_evidence,
        )
        yield row


def write_tsv(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        raise RuntimeError(f"refusing to write empty TSV: {path}")
    fields = list(rows[0])
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, delimiter="\t", lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    key: str(value).lower() if isinstance(value, bool) else value
                    for key, value in row.items()
                }
            )


def metric_for(source: dict[str, str]) -> RouteMetric:
    return RouteMetric(
        source_family=source["family"],
        publisher=source["publisher"],
        source_route=source["route"],
        route_schema=source["schema"],
        language=source["language"],
        rights_note=source["rights"],
    )


def update_metric(metric: RouteMetric, records: list[list[Atom]]) -> None:
    flat = [atom for record in records for atom in record]
    metric.raw_assertions += len(flat)
    metric.deinflated_assertions += sum(atom.counts_as_assertion for atom in flat)
    metric.strict_assertions += sum(
        atom.counts_as_assertion and atom.descriptor_class == "STRICT_FLAVOR" for atom in flat
    )
    metric.broad_assertions += sum(
        atom.counts_as_assertion and atom.descriptor_class == "BROAD_SENSORY" for atom in flat
    )
    metric.gold_assertions += sum(
        atom.counts_as_assertion and atom.collection_tier == "GOLD" for atom in flat
    )
    metric.silver_assertions += sum(
        atom.counts_as_assertion and atom.collection_tier == "SILVER" for atom in flat
    )
    metric.bronze_assertions += sum(
        atom.counts_as_assertion and atom.collection_tier == "BRONZE" for atom in flat
    )
    metric.descriptor_bearing_records += len(records)
    metric.record_ids.update(atom.effective_record_id for atom in flat)
    if flat:
        metric.descriptor_bearing_artifacts += 1
        metric.disposition = "POSITIVE_ROUTE_SCALED_OR_EXHAUSTED"


def public_metric_rows(metrics: list[RouteMetric]) -> list[dict[str, object]]:
    fields = (
        "source_family",
        "publisher",
        "source_route",
        "route_schema",
        "language",
        "artifacts_inspected",
        "descriptor_bearing_artifacts",
        "descriptor_bearing_records",
        "raw_assertions",
        "deinflated_assertions",
        "strict_assertions",
        "broad_assertions",
        "gold_assertions",
        "silver_assertions",
        "bronze_assertions",
        "duplicate_losses",
        "non_descriptor_losses",
        "analyst_equivalent_minutes",
        "automated_runtime_seconds",
        "rights_note",
        "disposition",
        "continuation_cursor",
        "basis",
    )
    result: list[dict[str, object]] = []
    for metric in metrics:
        row = {name: getattr(metric, name) for name in fields}
        row["automated_runtime_seconds"] = (
            "NA_ROUTE_TIMING_NOT_INSTRUMENTED;"
            "BATCH_RUNTIME_REPORTED_IN_PUBLIC_MANIFEST"
        )
        result.append(row)
    return result


def merged_milestone_snapshot(
    atoms: list[Atom],
    *,
    continuation: str,
    total_artifacts: int,
    assertion_losses: int,
    record_losses: int,
    restricted_path: Path,
    artifact_path: Path,
) -> dict[str, object]:
    """Build the immutable 10k checkpoint from the baseline plus Batch 2."""

    baseline_path = (
        ROOT / "db" / "data" / "current" / "CANONICAL_DESCRIPTOR_ASSERTION_LEDGER.tsv"
    )
    with baseline_path.open(encoding="utf-8", newline="") as handle:
        baseline_rows = [
            row
            for row in csv.DictReader(handle, delimiter="\t")
            if row["source_dataset_id"]
            != "professional-descriptor-scaleup-batch2-public-safe"
        ]
    baseline_assertions = [row for row in baseline_rows if row["counts_as_assertion"] == "true"]
    baseline_record_unique = [
        row
        for row in baseline_assertions
        if row["counts_as_record_unique_descriptor"] == "true"
    ]
    accepted_atoms = [atom for atom in atoms if atom.counts_as_assertion]
    record_unique_atoms = [
        atom for atom in accepted_atoms if atom.counts_as_record_unique_descriptor
    ]

    def merged_counter(baseline_field: str, atom_field: str) -> Counter[str]:
        result: Counter[str] = Counter(
            row[baseline_field] or "UNREPORTED" for row in baseline_assertions
        )
        result.update(getattr(atom, atom_field) or "UNREPORTED" for atom in accepted_atoms)
        return result

    family_counts = merged_counter("source_family_id", "source_family")
    year_counts = merged_counter("edition_year", "edition_year")
    rights_counts = merged_counter("rights_state", "rights_state")
    tier_counts = merged_counter("evidence_tier", "evidence_tier")
    class_counts = merged_counter("descriptor_class", "descriptor_class")
    collection_counts: Counter[str] = Counter()
    for row in baseline_assertions:
        collection_counts.update(
            ["GOLD" if row["evidence_tier"] in {"P1", "P2"} else "SILVER"]
        )
    collection_counts.update(atom.collection_tier for atom in accepted_atoms)

    preparation_counts: Counter[str] = Counter(
        row["preparation_service_id"] or "UNRESOLVED" for row in baseline_assertions
    )
    preparation_counts.update(
        atom.preparation_service or "UNREPORTED" for atom in accepted_atoms
    )
    roast_counts: Counter[str] = Counter(
        "UNREPORTED" if not row["source_native_roast_value"] else "DIRECT_EXPLICIT"
        for row in baseline_assertions
    )
    roast_counts.update(
        "UNREPORTED"
        if atom.roast_evidence in {"", "UNREPORTED", "SOURCE_UNKNOWN"}
        else "DIRECT_EXPLICIT"
        for atom in accepted_atoms
    )

    baseline_native_hashes = {
        row["atomic_source_text_sha256"]
        for row in baseline_assertions
        if row["atomic_source_text_sha256"]
    }
    new_native_hashes = {
        sha256_bytes(atom.atomic_text.encode("utf-8")) for atom in accepted_atoms
    }
    provisional_counts = Counter(atom.provisional_form_id for atom in accepted_atoms)
    support_bands: Counter[str] = Counter()
    for support in provisional_counts.values():
        if support == 1:
            support_bands["1"] += 1
        elif support <= 4:
            support_bands["2-4"] += 1
        elif support <= 9:
            support_bands["5-9"] += 1
        elif support <= 24:
            support_bands["10-24"] += 1
        elif support <= 49:
            support_bands["25-49"] += 1
        else:
            support_bands["50+"] += 1

    record_forms: dict[str, set[str]] = defaultdict(set)
    record_families: dict[str, str] = {}
    record_years: dict[str, str] = {}
    for atom in record_unique_atoms:
        record_forms[atom.effective_record_id].add(atom.provisional_form_id)
        record_families[atom.effective_record_id] = atom.source_family
        record_years[atom.effective_record_id] = atom.edition_year or "UNREPORTED"
    pair_support: Counter[tuple[str, str]] = Counter()
    pair_families: dict[tuple[str, str], set[str]] = defaultdict(set)
    pair_years: dict[tuple[str, str], set[str]] = defaultdict(set)
    pair_events = 0
    multi_target_records = 0
    for record_id, forms in record_forms.items():
        ordered = sorted(forms)
        if len(ordered) >= 2:
            multi_target_records += 1
        for left_index, left in enumerate(ordered):
            for right in ordered[left_index + 1 :]:
                pair = (left, right)
                pair_support[pair] += 1
                pair_families[pair].add(record_families[record_id])
                pair_years[pair].add(record_years[record_id])
                pair_events += 1

    total = len(baseline_assertions) + len(accepted_atoms)
    hhi = sum((count / total) ** 2 for count in family_counts.values()) if total else 0.0
    record_ids = {row["effective_record_id"] for row in baseline_assertions}
    record_ids.update(atom.effective_record_id for atom in accepted_atoms)
    judge_observations = {
        (atom.effective_record_id, atom.judge_observation_id)
        for atom in accepted_atoms
        if atom.publication_layer == "JUDGE_LEVEL_OBSERVATION" and atom.judge_observation_id
    }

    return {
        "schema": "professional-descriptor-candidate-milestone-v1",
        "batch_id": BATCH_ID,
        "milestone_role": "HEALTHY_CANDIDATE_CORPUS_MILESTONE",
        "stop_at_10000": False,
        "continue_after_10000": True,
        "hard_stop_assertion_count": HARD_STOP_TOTAL,
        "stop_rule": "FIRST_COMPLETE_RECORD_BOUNDARY_AT_OR_ABOVE_20000",
        "milestone_assertion_count": total,
        "raw_segmented_assertion_count": len(baseline_rows) + len(atoms),
        "record_unique_assertion_count": len(baseline_record_unique) + len(record_unique_atoms),
        "effective_record_count": len(record_ids),
        "judge_observation_count": len(judge_observations),
        "source_native_form_count": len(baseline_native_hashes | new_native_hashes),
        "provisional_normalized_form_count": len(provisional_counts),
        "provisional_mapping_coverage_count": len(accepted_atoms),
        "provisional_mapping_coverage_rate": len(accepted_atoms) / total if total else 0.0,
        "unmapped_baseline_assertion_count": len(baseline_assertions),
        "collection_tier_distribution": dict(sorted(collection_counts.items())),
        "evidence_tier_distribution": dict(sorted(tier_counts.items())),
        "descriptor_class_distribution": dict(sorted(class_counts.items())),
        "source_family_distribution": dict(sorted(family_counts.items())),
        "source_family_count": len(family_counts),
        "largest_source_family_share": max(family_counts.values()) / total if total else 0.0,
        "source_family_hhi": hhi,
        "year_distribution": dict(sorted(year_counts.items())),
        "preparation_distribution": dict(sorted(preparation_counts.items())),
        "roast_evidence_distribution": dict(sorted(roast_counts.items())),
        "rights_distribution": dict(sorted(rights_counts.items())),
        "duplicate_statistics": {
            "assertion_duplicate_losses": 1 + assertion_losses,
            "record_level_repeat_losses": 2 + record_losses,
        },
        "support_band_distribution": dict(sorted(support_bands.items())),
        "pair_statistics": {
            "effective_records_with_multiple_provisional_targets": multi_target_records,
            "provisional_normalized_pair_event_count": pair_events,
            "unique_provisional_normalized_pair_count": len(pair_support),
            "pairs_supported_by_multiple_source_families": sum(
                len(families) >= 2 for families in pair_families.values()
            ),
            "pairs_supported_by_multiple_years": sum(
                len(years) >= 2 for years in pair_years.values()
            ),
        },
        "artifact_count": total_artifacts,
        "exact_continuation_cursor": continuation,
        "public_safe_sidecar_sha256": sha256_file(
            PUBLIC_DIR / "PUBLIC_SAFE_ASSERTION_SIDECAR.tsv"
        ),
        "restricted_assertion_ledger_sha256": sha256_file(restricted_path),
        "restricted_raw_artifact_receipt_sha256": sha256_file(artifact_path),
        "model_eligible_assertion_count": 0,
        "rights_and_review_warning": "CANDIDATE_CORPUS_ONLY_NOT_MODEL_READY",
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--restricted-root", type=Path, required=True)
    parser.add_argument("--starting-count", type=int, default=157)
    parser.add_argument("--target-total", type=int, default=TARGET_TOTAL)
    parser.add_argument("--offline", action="store_true")
    parser.add_argument("--request-delay", type=float, default=0.65)
    args = parser.parse_args()
    started = time.monotonic()
    restricted = args.restricted_root / "professional_descriptor_batch2"
    raw_dir = restricted / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)

    robots_path = raw_dir / "coe-robots.txt"
    fetch(
        "https://farmdirectory.cupofexcellence.org/robots.txt",
        robots_path,
        offline=args.offline,
    )
    robots_text = robots_path.read_text(encoding="utf-8", errors="replace")
    if re.search(r"(?im)^\s*Disallow:\s*/\s*$", robots_text):
        raise RuntimeError("Cup of Excellence robots policy disallows crawling")

    all_atoms: list[Atom] = []
    metrics: list[RouteMetric] = []
    artifact_receipts: list[dict[str, object]] = []
    total_artifacts = 1

    parser_by_key = {
        "zenodo-q-grader-workbook": parse_zenodo,
        "project-origin-current-lots": parse_project_origin,
        "india-fine-cup-2019": parse_india_pdf,
        "isla-hawaii-2021": parse_isla_pdf,
        "fta-green-offering": parse_generic_cupping_pdf,
        "melbourne-coffee-merchants-2024": parse_generic_cupping_pdf,
        "sheba-haraz-2021": parse_sheba,
    }

    for key, source in STATIC_SOURCES.items():
        source_path = raw_dir / source["filename"]
        route_started = time.monotonic()
        try:
            downloaded = fetch(source["url"], source_path, offline=args.offline)
        except RuntimeError as exc:
            metric = metric_for(source)
            metric.artifacts_inspected = 0
            metric.disposition = "BLOCKED_PUBLIC_ACCESS"
            metric.continuation_cursor = f"RETRY:{source['url']}"
            metric.basis = str(exc)
            metrics.append(metric)
            print(f"STATIC_BLOCKED key={key} error={exc}", flush=True)
            continue
        total_artifacts += 1 if downloaded or source_path.is_file() else 0
        metric = metric_for(source)
        metric.artifacts_inspected = 1
        records = parser_by_key[key](source, source_path)
        flat = [atom for record in records for atom in record]
        assertion_losses, _ = apply_deinflation(flat)
        metric.duplicate_losses = assertion_losses
        update_metric(metric, records)
        metric.automated_runtime_seconds = f"{time.monotonic() - route_started:.6f}"
        metric.continuation_cursor = "EXHAUSTED_SINGLE_ARTIFACT_ROUTE"
        if not flat:
            metric.disposition = "ZERO_YIELD_PROBE_STOPPED"
        metrics.append(metric)
        all_atoms.extend(flat)
        artifact_receipts.append(
            {
                "source_route": source["route"],
                "source_url": source["url"],
                "restricted_relative_path": source_path.relative_to(restricted).as_posix(),
                "sha256": sha256_file(source_path),
                "byte_count": source_path.stat().st_size,
            }
        )
        print(
            f"STATIC_COMPLETE key={key} records={len(records)} assertions={sum(a.counts_as_assertion for a in flat)}",
            flush=True,
        )

    existing_total = args.starting_count + sum(atom.counts_as_assertion for atom in all_atoms)
    coe_links: set[str] = set()
    search_artifacts = 0
    for query in COE_SEARCH_QUERIES:
        query_key = re.sub(r"[^a-z0-9]+", "-", query.casefold()).strip("-")
        first_path = raw_dir / "coe-search" / f"{query_key}-page-001.html"
        fetch(coe_search_url(query, 1), first_path, offline=args.offline, delay=args.request_delay)
        search_artifacts += 1
        total_artifacts += 1
        payload = first_path.read_text(encoding="utf-8", errors="replace")
        links, page_count = coe_search_links(payload)
        coe_links.update(links)
        for page_number in range(2, page_count + 1):
            page_path = raw_dir / "coe-search" / f"{query_key}-page-{page_number:03d}.html"
            fetch(
                coe_search_url(query, page_number),
                page_path,
                offline=args.offline,
                delay=args.request_delay,
            )
            search_artifacts += 1
            total_artifacts += 1
            page_links, _ = coe_search_links(
                page_path.read_text(encoding="utf-8", errors="replace")
            )
            coe_links.update(page_links)
        print(
            f"COE_DISCOVERY query={query_key} pages={page_count} unique_links={len(coe_links)}",
            flush=True,
        )

    coe_metrics = {
        "route.coe.explicit-top-jury-detail": RouteMetric(
            source_family=COE["family"],
            publisher=COE["publisher"],
            source_route="route.coe.explicit-top-jury-detail",
            route_schema="schema.coe.explicit-top-jury-description.v2",
            language="en",
            rights_note="PENDING",
        ),
        "route.coe.generic-official-sensory-detail": RouteMetric(
            source_family=COE["family"],
            publisher=COE["publisher"],
            source_route="route.coe.generic-official-sensory-detail",
            route_schema="schema.coe.structured-official-sensory-fields.v2",
            language="en",
            rights_note="UNKNOWN",
        ),
    }
    coe_started = time.monotonic()
    continuation = "EXHAUSTED_DISCOVERED_DESCRIPTOR_SEARCH_LINKS"
    for index, source_url in enumerate(sorted(coe_links), 1):
        if existing_total >= args.target_total:
            continuation = f"detail-index={index};url={source_url}"
            break
        if total_artifacts >= MAX_ARTIFACTS:
            continuation = f"MAX_ARTIFACTS_REACHED;detail-index={index};url={source_url}"
            break
        path_key = hashlib.sha256(source_url.encode("utf-8")).hexdigest()[:24]
        detail_path = raw_dir / "coe-detail" / f"{path_key}.html"
        try:
            fetch(source_url, detail_path, offline=args.offline, delay=args.request_delay)
        except RuntimeError as exc:
            print(f"COE_DETAIL_BLOCKED index={index} url={source_url} error={exc}", flush=True)
            continuation = f"retry-detail-index={index};url={source_url}"
            continue
        total_artifacts += 1
        atoms = parse_coe_detail(detail_path, source_url)
        route = atoms[0].source_route if atoms else "route.coe.generic-official-sensory-detail"
        metric = coe_metrics[route]
        metric.artifacts_inspected += 1
        if atoms:
            apply_deinflation(atoms)
            metric.duplicate_losses += suppress_baseline_publication_overlap(
                atoms, source_url
            )
            update_metric(metric, [atoms])
            all_atoms.extend(atoms)
            existing_total += sum(atom.counts_as_assertion for atom in atoms)
            artifact_receipts.append(
                {
                    "source_route": route,
                    "source_url": source_url,
                    "restricted_relative_path": detail_path.relative_to(restricted).as_posix(),
                    "sha256": sha256_file(detail_path),
                    "byte_count": detail_path.stat().st_size,
                }
            )
        if index % 25 == 0:
            print(
                f"COE_SCALE inspected={index} staged_total={existing_total} target={args.target_total}",
                flush=True,
            )

    # Continue deterministically through the same official directory when the
    # descriptor-search slices are exhausted. Previously seen search URLs are
    # skipped; each newly accepted detail remains one complete record boundary.
    if (
        existing_total < args.target_total
        and continuation == "EXHAUSTED_DISCOVERED_DESCRIPTOR_SEARCH_LINKS"
        and total_artifacts < MAX_ARTIFACTS
    ):
        seen_urls = set(coe_links)
        archive_exhausted = True
        for page_number in range(1, MAX_COE_ARCHIVE_PAGES + 1):
            if existing_total >= args.target_total or total_artifacts >= MAX_ARTIFACTS:
                archive_exhausted = False
                break
            archive_path = raw_dir / "coe-archive" / f"listings-page-{page_number:03d}.html"
            try:
                fetch(
                    coe_archive_url(page_number),
                    archive_path,
                    offline=args.offline,
                    delay=args.request_delay,
                )
            except RuntimeError as exc:
                continuation = f"retry-archive-page={page_number};error={exc}"
                archive_exhausted = False
                break
            total_artifacts += 1
            archive_urls = sorted(
                coe_archive_links(
                    archive_path.read_text(encoding="utf-8", errors="replace")
                )
            )
            if not archive_urls:
                continuation = f"archive-page={page_number};NO_LISTING_LINKS"
                archive_exhausted = False
                break
            for detail_index, source_url in enumerate(archive_urls, 1):
                if source_url in seen_urls:
                    continue
                seen_urls.add(source_url)
                if existing_total >= args.target_total:
                    continuation = (
                        f"archive-page={page_number};detail-index={detail_index};url={source_url}"
                    )
                    archive_exhausted = False
                    break
                if total_artifacts >= MAX_ARTIFACTS:
                    continuation = (
                        "MAX_ARTIFACTS_REACHED;"
                        f"archive-page={page_number};detail-index={detail_index};url={source_url}"
                    )
                    archive_exhausted = False
                    break
                path_key = hashlib.sha256(source_url.encode("utf-8")).hexdigest()[:24]
                detail_path = raw_dir / "coe-detail" / f"{path_key}.html"
                try:
                    fetch(
                        source_url,
                        detail_path,
                        offline=args.offline,
                        delay=args.request_delay,
                    )
                except RuntimeError as exc:
                    print(
                        f"COE_ARCHIVE_DETAIL_BLOCKED page={page_number} "
                        f"index={detail_index} url={source_url} error={exc}",
                        flush=True,
                    )
                    continuation = (
                        f"retry-archive-page={page_number};detail-index={detail_index};url={source_url}"
                    )
                    continue
                total_artifacts += 1
                atoms = parse_coe_detail(detail_path, source_url)
                route = (
                    atoms[0].source_route
                    if atoms
                    else "route.coe.generic-official-sensory-detail"
                )
                metric = coe_metrics[route]
                metric.artifacts_inspected += 1
                if atoms:
                    apply_deinflation(atoms)
                    metric.duplicate_losses += suppress_baseline_publication_overlap(
                        atoms, source_url
                    )
                    update_metric(metric, [atoms])
                    all_atoms.extend(atoms)
                    existing_total += sum(atom.counts_as_assertion for atom in atoms)
                    artifact_receipts.append(
                        {
                            "source_route": route,
                            "source_url": source_url,
                            "restricted_relative_path": detail_path.relative_to(restricted).as_posix(),
                            "sha256": sha256_file(detail_path),
                            "byte_count": detail_path.stat().st_size,
                        }
                    )
            if not archive_exhausted:
                break
            if page_number % 10 == 0:
                print(
                    f"COE_ARCHIVE_SCALE page={page_number} staged_total={existing_total} "
                    f"target={args.target_total}",
                    flush=True,
                )
        if archive_exhausted:
            continuation = "EXHAUSTED_COE_ARCHIVE"

    for metric in coe_metrics.values():
        metric.automated_runtime_seconds = f"{time.monotonic() - coe_started:.6f}"
        metric.continuation_cursor = continuation
        if metric.deinflated_assertions:
            metric.disposition = (
                "POSITIVE_ROUTE_CHECKPOINT_TARGET_REACHED"
                if existing_total >= args.target_total
                else "POSITIVE_ROUTE_NOT_EXHAUSTED"
            )
        elif metric.artifacts_inspected:
            metric.disposition = "ZERO_YIELD_INSPECTED_ROUTE_STRATUM"
        metrics.append(metric)

    final_assertion_losses, final_record_losses = apply_deinflation(all_atoms)
    cross_batch_publication_losses = sum(
        atom.source_url in BASELINE_PUBLICATION_OVERLAP_URLS
        and atom.descriptor_class != "NON_DESCRIPTOR"
        for atom in all_atoms
    )
    final_total = args.starting_count + sum(atom.counts_as_assertion for atom in all_atoms)
    if final_total < args.target_total and continuation in {
        "EXHAUSTED_DISCOVERED_DESCRIPTOR_SEARCH_LINKS",
        "EXHAUSTED_COE_ARCHIVE",
    }:
        phase_status = "CANDIDATE_CORPUS_PROGRESS_CHECKPOINT"
    elif final_total >= HARD_STOP_TOTAL:
        phase_status = "CANDIDATE_CORPUS_20000_REACHED_DISTRIBUTION_GAPS"
    elif final_total >= MILESTONE_TOTAL:
        phase_status = "CANDIDATE_CORPUS_10000_REACHED_DISTRIBUTION_GAPS"
    else:
        phase_status = "CANDIDATE_CORPUS_PROGRESS_CHECKPOINT"

    safe = list(safe_rows(all_atoms))
    restricted_data = list(restricted_rows(all_atoms))
    public_path = PUBLIC_DIR / "PUBLIC_SAFE_ASSERTION_SIDECAR.tsv"
    restricted_path = restricted / "PROFESSIONAL_ASSERTIONS_RESTRICTED.tsv"
    write_tsv(public_path, safe)
    write_tsv(restricted_path, restricted_data)

    route_metrics = metrics + [RouteMetric(**row) for row in HISTORICAL_PROBES]
    route_rows = public_metric_rows(route_metrics)
    route_path = PUBLIC_DIR / "SOURCE_ROUTE_PROBE_AND_YIELD.tsv"
    write_tsv(route_path, route_rows)

    artifact_path = restricted / "RAW_ARTIFACT_RECEIPT.tsv"
    write_tsv(artifact_path, artifact_receipts)
    artifact_root_hash = sha256_file(artifact_path)
    restricted_receipt = {
        "batch_id": BATCH_ID,
        "storage_class": "OWNER_CONTROLLED_RESTRICTED_NON_GIT",
        "restricted_assertion_ledger": {
            "path": restricted_path.name,
            "sha256": sha256_file(restricted_path),
            "row_count": len(restricted_data),
        },
        "raw_artifact_receipt": {
            "path": artifact_path.name,
            "sha256": artifact_root_hash,
            "row_count": len(artifact_receipts),
        },
    }
    (restricted / "RESTRICTED_CHECKPOINT.json").write_text(
        json.dumps(restricted_receipt, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )

    milestone_path = PUBLIC_DIR / "CANDIDATE_MILESTONE_10000.json"
    restricted_milestone_path = restricted / "MILESTONE_10000_RESTRICTED_CHECKPOINT.json"
    if final_total >= MILESTONE_TOTAL and not milestone_path.exists():
        milestone = merged_milestone_snapshot(
            all_atoms,
            continuation=continuation,
            total_artifacts=total_artifacts,
            assertion_losses=final_assertion_losses,
            record_losses=final_record_losses,
            restricted_path=restricted_path,
            artifact_path=artifact_path,
        )
        milestone_path.write_text(
            json.dumps(milestone, indent=2, sort_keys=True) + "\n", encoding="utf-8"
        )
        restricted_milestone_path.write_text(
            json.dumps(
                {
                    "schema": "professional-descriptor-restricted-milestone-v1",
                    "batch_id": BATCH_ID,
                    "exact_continuation_cursor": continuation,
                    "public_milestone_sha256": sha256_file(milestone_path),
                    "restricted_assertion_ledger_sha256": sha256_file(restricted_path),
                    "restricted_raw_artifact_receipt_sha256": sha256_file(artifact_path),
                },
                indent=2,
                sort_keys=True,
            )
            + "\n",
            encoding="utf-8",
        )
        print(f"MILESTONE_10000_CREATED cursor={continuation}", flush=True)
    elif milestone_path.exists():
        print("MILESTONE_10000_PRESERVED", flush=True)

    if final_total >= HARD_STOP_TOTAL and milestone_path.exists():
        corrected_total = args.starting_count
        prefix_end = 0
        completed_source_url = ""
        for _, group_iter in itertools.groupby(
            enumerate(all_atoms), key=lambda item: item[1].effective_record_id
        ):
            group = list(group_iter)
            corrected_total += sum(atom.counts_as_assertion for _, atom in group)
            prefix_end = group[-1][0] + 1
            completed_source_url = group[0][1].source_url
            if corrected_total >= MILESTONE_TOTAL:
                break
        prefix_atoms = all_atoms[:prefix_end]
        next_atom = all_atoms[prefix_end] if prefix_end < len(all_atoms) else None
        sorted_search_links = sorted(coe_links)
        if next_atom and next_atom.source_url in coe_links:
            next_cursor = (
                f"detail-index={sorted_search_links.index(next_atom.source_url) + 1};"
                f"url={next_atom.source_url}"
            )
        elif next_atom:
            next_cursor = f"next-effective-record={next_atom.effective_record_id};url={next_atom.source_url}"
        else:
            next_cursor = "EOF"
        milestone_artifact_count = (
            1
            + search_artifacts
            + len({atom.artifact_sha256 for atom in prefix_atoms})
        )
        reconciliation = merged_milestone_snapshot(
            prefix_atoms,
            continuation=next_cursor,
            total_artifacts=milestone_artifact_count,
            assertion_losses=sum(not atom.counts_as_assertion for atom in prefix_atoms),
            record_losses=sum(
                atom.counts_as_assertion
                and not atom.counts_as_record_unique_descriptor
                for atom in prefix_atoms
            ),
            restricted_path=restricted_path,
            artifact_path=artifact_path,
        )
        original_milestone = json.loads(milestone_path.read_text(encoding="utf-8"))
        reconciliation.update(
            schema="professional-descriptor-candidate-milestone-reconciliation-v1",
            artifact_count=original_milestone["artifact_count"] + 1,
            original_milestone_receipt_sha256=sha256_file(milestone_path),
            original_acquisition_local_milestone_assertion_count=original_milestone[
                "milestone_assertion_count"
            ],
            original_acquisition_local_cursor=original_milestone[
                "exact_continuation_cursor"
            ],
            cross_batch_publication_layer_duplicate_losses=18,
            merged_count_at_original_cursor=(
                original_milestone["milestone_assertion_count"] - 18
            ),
            corrected_first_complete_record_boundary_count=corrected_total,
            corrected_completed_source_url=completed_source_url,
            corrected_exact_continuation_cursor=next_cursor,
            public_safe_prefix_row_count=prefix_end,
            public_safe_prefix_last_assertion_id=(
                prefix_atoms[-1].descriptor_assertion_id if prefix_atoms else ""
            ),
            full_public_safe_sidecar_sha256=sha256_file(public_path),
            reconciliation_reason=(
                "BATCH1_PUBLICATION_REAPPEARED_IN_BATCH2_AND_WAS_RETAINED_AS_"
                "FALSE_COUNT_AUDIT_ROWS"
            ),
        )
        (PUBLIC_DIR / "CANDIDATE_MILESTONE_10000_RECONCILIATION.json").write_text(
            json.dumps(reconciliation, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )

    family_counts = Counter(
        atom.source_family for atom in all_atoms if atom.counts_as_assertion
    )
    public_manifest = {
        "schema": "professional-descriptor-batch2-public-safe-v1",
        "batch_id": BATCH_ID,
        "generated_date": BATCH_DATE,
        "phase_status": phase_status,
        "starting_deinflated_candidate_count": args.starting_count,
        "net_new_deinflated_candidate_count": sum(
            atom.counts_as_assertion for atom in all_atoms
        ),
        "final_deinflated_candidate_count": final_total,
        "remaining_to_10000": max(MILESTONE_TOTAL - final_total, 0),
        "candidate_corpus_10000_reached": final_total >= MILESTONE_TOTAL,
        "target_10000_role": "HEALTHY_CANDIDATE_CORPUS_MILESTONE",
        "stop_at_10000": False,
        "continue_after_10000": True,
        "hard_stop_assertion_count": HARD_STOP_TOTAL,
        "stop_at_first_complete_record_boundary_at_or_above_20000": True,
        "raw_segmented_new_assertion_count": len(all_atoms),
        "record_unique_new_assertion_count": sum(
            atom.counts_as_record_unique_descriptor for atom in all_atoms
        ),
        "assertion_duplicate_losses": final_assertion_losses
        + cross_batch_publication_losses,
        "cross_batch_publication_layer_duplicate_losses": cross_batch_publication_losses,
        "record_level_repeat_losses": final_record_losses,
        "source_family_count_new": len(family_counts),
        "source_route_count_inspected": len(route_metrics),
        "positive_source_family_count": len(
            {
                metric.source_family
                for metric in route_metrics
                if metric.deinflated_assertions > 0
            }
        ),
        "largest_new_source_family_share": (
            max(family_counts.values()) / sum(family_counts.values()) if family_counts else 0
        ),
        "collection_tier_distribution": Counter(
            atom.collection_tier for atom in all_atoms if atom.counts_as_assertion
        ),
        "evidence_tier_distribution": Counter(
            atom.evidence_tier for atom in all_atoms if atom.counts_as_assertion
        ),
        "descriptor_class_distribution": Counter(
            atom.descriptor_class for atom in all_atoms if atom.counts_as_assertion
        ),
        "rights_distribution": Counter(
            atom.rights_state for atom in all_atoms if atom.counts_as_assertion
        ),
        "provisional_form_count": len(
            {atom.provisional_form_id for atom in all_atoms if atom.counts_as_assertion}
        ),
        "provisional_mapping_coverage_rate": 1.0,
        "artifact_count": total_artifacts,
        "automated_live_acquisition_runtime_seconds": (
            OBSERVED_LIVE_ACQUISITION_RUNTIME_SECONDS
        ),
        "analyst_equivalent_minutes": "NA_AUTOMATED_ACQUISITION_NO_MANUAL_TIMING",
        "continuation_cursor": continuation,
        "restricted_checkpoint_receipt_sha256": sha256_file(
            restricted / "RESTRICTED_CHECKPOINT.json"
        ),
        "files": [],
    }
    for path in (
        public_path,
        route_path,
        milestone_path,
        PUBLIC_DIR / "CANDIDATE_MILESTONE_10000_RECONCILIATION.json",
    ):
        public_manifest["files"].append(
            {
                "path": path.relative_to(PUBLIC_DIR).as_posix(),
                "sha256": sha256_file(path),
                "byte_count": path.stat().st_size,
                "row_count": (
                    max(sum(1 for _ in path.open(encoding="utf-8")) - 1, 0)
                    if path.suffix == ".tsv"
                    else "NA_NOT_TABULAR"
                ),
            }
        )
    manifest_path = PUBLIC_DIR / "BATCH2_PUBLIC_MANIFEST.json"
    manifest_path.write_text(
        json.dumps(public_manifest, indent=2, sort_keys=True, default=dict) + "\n",
        encoding="utf-8",
    )
    checksum_lines = [
        f"{sha256_file(path)}  {path.name}"
        for path in sorted(PUBLIC_DIR.iterdir())
        if path.is_file() and path.name != "SHA256SUMS"
    ]
    (PUBLIC_DIR / "SHA256SUMS").write_text("\n".join(checksum_lines) + "\n", encoding="utf-8")

    elapsed = time.monotonic() - started
    print(f"PHASE_STATUS={phase_status}")
    print(f"NET_NEW_DEINFLATED_CANDIDATE_COUNT={public_manifest['net_new_deinflated_candidate_count']}")
    print(f"FINAL_DEINFLATED_CANDIDATE_COUNT={final_total}")
    print(f"SOURCE_FAMILY_COUNT_NEW={len(family_counts)}")
    print(f"POSITIVE_SOURCE_FAMILY_COUNT={public_manifest['positive_source_family_count']}")
    print(f"ACQUIRED_ARTIFACT_COUNT={total_artifacts}")
    print(f"AUTOMATED_RUNTIME_SECONDS={elapsed:.6f}")
    print(f"NEXT_CONTINUATION_CURSOR={continuation}")


if __name__ == "__main__":
    main()

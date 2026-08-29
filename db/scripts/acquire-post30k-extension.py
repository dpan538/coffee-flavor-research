#!/usr/bin/env python3
"""Continue the exact 30k cursor into isolated post-30k acquisition staging.

Source-native text and downloaded artifacts remain under ``--restricted-root``.
Only hashes, stable identifiers, locators, rights states, and aggregate receipts
are committed.  The run stops at the first complete record boundary at or above
40,000 total de-inflated candidate source assertions.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import importlib.util
import io
import json
import re
import sys
import zipfile
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
PUBLIC = ROOT / "db" / "data" / "post30k-extension-staging"
FROZEN_20K = Path("/private/tmp/round3l-acquisition/professional_descriptor_batch2")
POST20K = Path("/private/tmp/coffee-flavor-round3m-post20k/post20k_extension")
FROZEN_COUNT = 30_010
TARGET_TOTAL = 40_000
START_PAGE = 75
START_INDEX = 2
START_URL = "https://farmdirectory.cupofexcellence.org/listing/2-don-dario-hacienda-san-isidro-labrador-costa-rica-2024-experimental/"
CURSOR_START = f"archive-page={START_PAGE};detail-index={START_INDEX};url={START_URL}"
EXTENSION_BATCH_ID = "professional-descriptor-post30k-extension-20260829"
PARSER_VERSION = "post30k.professional-descriptor-parser.v1"
ADAPTER_VERSION = "post30k.public-safe-adapter.v1"
GENERATED_AT = "2026-08-29T00:00:00Z"

MDPI_SUPPLEMENT_URL = (
    "https://mdpi-res.com/d_attachment/foods/foods-15-02756/"
    "article_deploy/foods-15-02756-s001.zip"
)
MDPI_XML_URL = (
    "https://mdpi-res.com/d_attachment/foods/foods-15-02756/"
    "article_deploy/foods-15-02756.xml"
)
MDPI_INNER_XLSX = "File S1. Sensory Evaluation Results.xlsx"


def load_post20_module():
    path = ROOT / "db" / "scripts" / "acquire-post20k-extension.py"
    spec = importlib.util.spec_from_file_location("post20k_acquisition", path)
    if spec is None or spec.loader is None:
        raise RuntimeError("cannot import post-20k acquisition primitives")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


P20 = load_post20_module()
B2 = P20.B2


def sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def write_tsv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        raise RuntimeError(f"refusing to write empty TSV: {path}")
    fields = list(rows[0])
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=fields,
            delimiter="\t",
            lineterminator="\n",
            extrasaction="raise",
        )
        writer.writeheader()
        for row in rows:
            writer.writerow({
                key: str(value).lower() if isinstance(value, bool) else value
                for key, value in row.items()
            })


MDPI_SOURCE = P20.source_definition(
    family="family.mdpi_certified_q_grader_storage_panel",
    route="route.mdpi.2026.certified-q-grader-storage-narratives",
    schema="schema.certified-q-grader.panel-aggregate-sensory-narrative.v1",
    publisher="MDPI Foods / Universidad Nacional Toribio Rodríguez de Mendoza",
    rights="AFFIRMATIVE",
    rights_basis="CC_BY_4_0_MDPI_ARTICLE_AND_SUPPLEMENT",
    evidence_tier="P2",
    collection_tier="GOLD",
)
MDPI_SOURCE["language"] = "es"


def parse_mdpi_supplement(path: Path) -> tuple[list[list[Any]], str]:
    with zipfile.ZipFile(path) as archive:
        if MDPI_INNER_XLSX not in archive.namelist():
            raise RuntimeError("MDPI Supplementary File S1 workbook is missing")
        payload = archive.read(MDPI_INNER_XLSX)
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    if "Descriptors" not in workbook.sheetnames:
        raise RuntimeError("MDPI descriptor narrative sheet is missing")
    sheet = workbook["Descriptors"]
    artifact_hash = sha256_file(path)
    inner_hash = sha256_bytes(payload)
    section_id = "unresolved-section"
    records: list[list[Any]] = []
    for row_index, values in enumerate(sheet.iter_rows(values_only=True), start=1):
        cells = list(values)
        if cells and isinstance(cells[0], str) and "Perfil sensorial" in cells[0]:
            section_id = B2.stable_id("mdpi-storage-section", cells[0])
            continue
        for block_index, (day_col, narrative_col) in enumerate(((0, 2), (3, 5), (6, 8)), start=1):
            if len(cells) <= narrative_col:
                continue
            narrative = cells[narrative_col]
            if not isinstance(narrative, str) or len(narrative.strip()) < 6:
                continue
            if "Descripción" in narrative:
                continue
            day = cells[day_col]
            day_id = str(day).strip() if day is not None else "unreported"
            effective = B2.stable_id(
                "effective-post30k", MDPI_SOURCE["route"], section_id, row_index, block_index, day_id
            )
            coffee = B2.stable_id("coffee-post30k", MDPI_SOURCE["family"], section_id)
            atoms = B2.make_atoms(
                source=MDPI_SOURCE,
                artifact_sha256=artifact_hash,
                source_url=MDPI_SUPPLEMENT_URL,
                source_locator=(
                    f"zip:{MDPI_INNER_XLSX}#sheet:Descriptors;row={row_index};"
                    f"block={block_index};day={day_id}"
                ),
                effective_record_id=effective,
                coffee_identity_id=coffee,
                edition_or_release="Foods 2026 15(15) 2756 Supplementary File S1",
                edition_year="2026",
                preparation_service="CERTIFIED_Q_ARABICA_GRADER_CUPPING",
                roast_evidence="MULTIPLE_GREEN_ROASTED_AND_GROUND_STORAGE_CONDITIONS_PER_ARTICLE",
                source_field_label="Descripción",
                raw_field_text=narrative,
                publication_layer="CERTIFIED_Q_GRADER_PANEL_AGGREGATED_NARRATIVE",
                provenance_state="SOURCE_AUDITED_CERTIFIED_Q_GRADER_PANEL_NARRATIVE",
            )
            records.append(atoms)
    if len(records) < 100:
        raise RuntimeError(f"MDPI descriptor record count drift: {len(records)}")
    return records, inner_hash


def static_non_coe_discovery() -> list[dict[str, Any]]:
    routes = [
        (
            "family.wiley_coffee_character_wheel_panel",
            "route.wiley.2023.character-wheel-supporting-data",
            "POSITIVE_LEXICON_ONLY_NOT_ROW_LEVEL_STAGED",
            "SEMI_TRAINED_COFFEE_PROFESSIONAL_PANEL_LEXICON_DOES_NOT PROVIDE_SAMPLE_LEVEL_OBSERVATION_MATRIX",
        ),
        (
            "family.dryad_uc_davis_black_coffee",
            "route.dryad.10.25338.B8993H.consumer-cata",
            "EXCLUDED_CONSUMER_NOT_PROFESSIONAL",
            "OPEN_ROW_LEVEL_CATA_DATA_ARE CONSUMER OBSERVATIONS_OUTSIDE_PROFESSIONAL_SCOPE",
        ),
        (
            "family.springer_metabolomics_ground_coffee",
            "route.springer.2020.metabolomics-sensory-scores",
            "ZERO_YIELD_SCORES_NO_SOURCE_NATIVE_DESCRIPTORS",
            "TRAINED_PANEL_SCORES_PRESENT_WITHOUT_ROW_LEVEL_DESCRIPTOR NARRATIVES",
        ),
        (
            "family.frontiers_wet_processing_panel",
            "route.frontiers.2019.wet-processing-trained-panel",
            "POSITIVE_AGGREGATE_ONLY_NOT_ROW_LEVEL_STAGED",
            "TRAINED_PANEL_CONFIRMED_BUT OPEN SUPPLEMENT DOES NOT EXPOSE SOURCE NATIVE NARRATIVE ROWS",
        ),
        (
            "family.researchsquare_specialty_quality",
            "route.researchsquare.2024.datasetcomplete",
            "BLOCKED_RIGHTS_AND_VERSION_STABILITY",
            "PREPRINT SUPPLEMENT DISCOVERED WITHOUT STABLE RIGHTS-CLEARABLE VERSIONED ROW CONTRACT",
        ),
    ]
    return [{
        "source_family_id": family,
        "publisher_id": B2.stable_id("publisher", family),
        "source_route_id": route,
        "route_schema": "schema.discovery-audit.v1",
        "discovery_lane": "NON_COE",
        "artifacts_inspected": 1,
        "descriptor_bearing_records": 0,
        "raw_assertion_count": 0,
        "deinflated_assertion_count": 0,
        "rights_state": "UNKNOWN",
        "disposition": disposition,
        "basis": basis.replace(" ", "_"),
    } for family, route, disposition, basis in routes]


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--restricted-root", type=Path, required=True)
    parser.add_argument("--offline", action="store_true")
    parser.add_argument("--request-delay", type=float, default=0.2)
    args = parser.parse_args()

    restricted = args.restricted_root / "post30k_extension"
    raw = restricted / "raw"
    raw.mkdir(parents=True, exist_ok=True)
    PUBLIC.mkdir(parents=True, exist_ok=True)

    frozen_rows = read_tsv(FROZEN_20K / "PROFESSIONAL_ASSERTIONS_RESTRICTED.tsv")
    post20_rows = read_tsv(POST20K / "POST20K_ASSERTIONS_RESTRICTED.tsv")
    prior_urls = {row["source_url"] for row in frozen_rows + post20_rows}
    if START_URL in prior_urls:
        raise RuntimeError("exact post-30k start cursor was already acquired")

    all_atoms: list[Any] = []
    cursor_by_record: dict[str, str] = {}
    artifact_rows: list[dict[str, Any]] = []
    route_rows: list[dict[str, Any]] = []

    mdpi_zip = raw / "non-coe" / "foods-15-02756-s001.zip"
    mdpi_xml = raw / "non-coe" / "foods-15-02756.xml"
    P20.fetch_retry(MDPI_SUPPLEMENT_URL, mdpi_zip, offline=args.offline, delay=args.request_delay)
    P20.fetch_retry(MDPI_XML_URL, mdpi_xml, offline=args.offline, delay=args.request_delay)
    xml_text = mdpi_xml.read_text(encoding="utf-8", errors="replace")
    if "creativecommons.org/licenses/by/4.0" not in xml_text.casefold():
        raise RuntimeError("MDPI CC BY 4.0 license locator drift")
    mdpi_records, inner_hash = parse_mdpi_supplement(mdpi_zip)
    mdpi_atoms = [atom for record in mdpi_records for atom in record]
    B2.apply_deinflation(mdpi_atoms)
    all_atoms.extend(mdpi_atoms)
    for atom in mdpi_atoms:
        cursor_by_record[atom.effective_record_id] = "noncoe-route=mdpi-certified-q-grader-storage;status=EXHAUSTED"
    route_rows.append({
        "source_family_id": MDPI_SOURCE["family"],
        "publisher_id": B2.stable_id("publisher", MDPI_SOURCE["family"]),
        "source_route_id": MDPI_SOURCE["route"],
        "route_schema": MDPI_SOURCE["schema"],
        "discovery_lane": "NON_COE",
        "artifacts_inspected": 2,
        "descriptor_bearing_records": len(mdpi_records),
        "raw_assertion_count": len(mdpi_atoms),
        "deinflated_assertion_count": sum(atom.counts_as_assertion for atom in mdpi_atoms),
        "rights_state": "AFFIRMATIVE_WITH_CONDITIONS",
        "disposition": "POSITIVE_RIGHTS_CLEARABLE_ROUTE_EXHAUSTED",
        "basis": "CC_BY_4_0_CERTIFIED_Q_GRADER_PANEL_NARRATIVES_SOURCE_LANGUAGE_PRESERVED",
    })
    for url, path in ((MDPI_SUPPLEMENT_URL, mdpi_zip), (MDPI_XML_URL, mdpi_xml)):
        artifact_rows.append({
            "source_route_id": MDPI_SOURCE["route"],
            "source_url": url,
            "restricted_relative_path": path.relative_to(restricted).as_posix(),
            "sha256": sha256_file(path),
            "byte_count": path.stat().st_size,
            "acquisition_cursor": "noncoe-route=mdpi-certified-q-grader-storage;status=EXHAUSTED",
        })
    artifact_rows.append({
        "source_route_id": MDPI_SOURCE["route"],
        "source_url": MDPI_SUPPLEMENT_URL + "#" + MDPI_INNER_XLSX,
        "restricted_relative_path": "restricted-archive-member://" + MDPI_INNER_XLSX,
        "sha256": inner_hash,
        "byte_count": "ARCHIVE_MEMBER",
        "acquisition_cursor": "noncoe-route=mdpi-certified-q-grader-storage;status=EXHAUSTED",
    })
    route_rows.extend(static_non_coe_discovery())

    accepted = sum(atom.counts_as_assertion for atom in all_atoms)
    coe_records: list[list[Any]] = []
    cursor_end = CURSOR_START
    blocked = False
    route_exhausted = False
    exact_cursor_validated = False
    for page_number in range(START_PAGE, B2.MAX_COE_ARCHIVE_PAGES + 1):
        archive_path = raw / "coe-archive" / f"listings-page-{page_number:03d}.html"
        try:
            P20.fetch_retry(B2.coe_archive_url(page_number), archive_path, offline=args.offline, delay=args.request_delay)
        except RuntimeError:
            cursor_end = f"retry-archive-page={page_number}"
            blocked = True
            break
        urls = sorted(B2.coe_archive_links(archive_path.read_text(encoding="utf-8", errors="replace")))
        artifact_rows.append({
            "source_route_id": "route.coe.post30k-archive-continuation",
            "source_url": B2.coe_archive_url(page_number),
            "restricted_relative_path": archive_path.relative_to(restricted).as_posix(),
            "sha256": sha256_file(archive_path),
            "byte_count": archive_path.stat().st_size,
            "acquisition_cursor": f"archive-page={page_number}",
        })
        if not urls:
            cursor_end = f"archive-page={page_number};NO_LISTING_LINKS"
            blocked = True
            break
        first_index = START_INDEX if page_number == START_PAGE else 1
        if page_number == START_PAGE:
            if len(urls) < START_INDEX or urls[START_INDEX - 1] != START_URL:
                raise RuntimeError("preserved post-30k cursor no longer matches archive page 75")
            exact_cursor_validated = True
        for detail_index, source_url in enumerate(urls, start=1):
            if detail_index < first_index:
                continue
            if source_url in prior_urls:
                continue
            if FROZEN_COUNT + accepted >= TARGET_TOTAL:
                cursor_end = f"archive-page={page_number};detail-index={detail_index};url={source_url}"
                break
            detail_path = raw / "coe-detail" / f"{hashlib.sha256(source_url.encode()).hexdigest()[:24]}.html"
            try:
                P20.fetch_retry(source_url, detail_path, offline=args.offline, delay=args.request_delay)
            except RuntimeError:
                cursor_end = f"retry-archive-page={page_number};detail-index={detail_index};url={source_url}"
                blocked = True
                break
            atoms = B2.parse_coe_detail(detail_path, source_url)
            B2.apply_deinflation(atoms)
            if atoms:
                coe_records.append(atoms)
                all_atoms.extend(atoms)
                accepted += sum(atom.counts_as_assertion for atom in atoms)
                cursor = f"archive-page={page_number};detail-index={detail_index};url={source_url}"
                for atom in atoms:
                    cursor_by_record[atom.effective_record_id] = cursor
                artifact_rows.append({
                    "source_route_id": atoms[0].source_route,
                    "source_url": source_url,
                    "restricted_relative_path": detail_path.relative_to(restricted).as_posix(),
                    "sha256": sha256_file(detail_path),
                    "byte_count": detail_path.stat().st_size,
                    "acquisition_cursor": cursor,
                })
            if coe_records and len(coe_records) % 50 == 0:
                print(
                    f"POST30K_COE_PROGRESS records={len(coe_records)} extension={accepted} total={FROZEN_COUNT + accepted}",
                    flush=True,
                )
        if blocked or FROZEN_COUNT + accepted >= TARGET_TOTAL:
            break
    else:
        route_exhausted = True
        cursor_end = "EXHAUSTED_COE_ARCHIVE"

    coe_atoms = [atom for record in coe_records for atom in record]
    route_rows.append({
        "source_family_id": B2.COE["family"],
        "publisher_id": B2.stable_id("publisher", B2.COE["family"]),
        "source_route_id": "route.coe.post30k-archive-continuation",
        "route_schema": "schema.coe.mixed-explicit-jury-and-generic.v2",
        "discovery_lane": "COE_CONTINUATION",
        "artifacts_inspected": len(coe_records),
        "descriptor_bearing_records": len(coe_records),
        "raw_assertion_count": len(coe_atoms),
        "deinflated_assertion_count": sum(atom.counts_as_assertion for atom in coe_atoms),
        "rights_state": "MIXED_PENDING_UNKNOWN",
        "disposition": "POSITIVE_ROUTE_CHECKPOINT_REACHED" if FROZEN_COUNT + accepted >= TARGET_TOTAL else "PARTIAL_OR_EXHAUSTED",
        "basis": "EXACT_30K_CURSOR_CONTINUATION_FIRST_COMPLETE_RECORD_BOUNDARY",
    })

    B2.apply_deinflation(all_atoms)
    accepted = sum(atom.counts_as_assertion for atom in all_atoms)
    raw_count = len(all_atoms)
    record_unique = sum(atom.counts_as_record_unique_descriptor for atom in all_atoms)
    effective_records = {atom.effective_record_id for atom in all_atoms if atom.counts_as_assertion}
    non_coe_atoms = [atom for atom in all_atoms if atom.source_family != B2.COE["family"]]
    coe_atoms = [atom for atom in all_atoms if atom.source_family == B2.COE["family"]]
    positive_non_coe = {atom.source_family for atom in non_coe_atoms if atom.counts_as_assertion}

    safe_rows: list[dict[str, Any]] = []
    for atom, safe in zip(all_atoms, B2.safe_rows(all_atoms)):
        safe_public = dict(safe)
        safe_public.pop("publisher", None)
        safe_public["publisher_id"] = B2.stable_id("publisher", atom.source_family)
        safe_rows.append({
            "extension_batch_id": EXTENSION_BATCH_ID,
            **safe_public,
            "source_field_label": "hash:sha256:" + hashlib.sha256(atom.source_field_label.encode()).hexdigest(),
            "parser_version": PARSER_VERSION,
            "adapter_version": ADAPTER_VERSION,
            "acquisition_cursor": cursor_by_record[atom.effective_record_id],
            "frozen_snapshot_version": "professional-descriptor-candidate-v1-30k",
            "frozen_snapshot_member": False,
        })
    restricted_rows = list(B2.restricted_rows(all_atoms))
    restricted_ledger = restricted / "POST30K_ASSERTIONS_RESTRICTED.tsv"
    restricted_artifacts = restricted / "POST30K_RAW_ARTIFACT_RECEIPT.tsv"
    B2.write_tsv(restricted_ledger, restricted_rows)
    write_tsv(restricted_artifacts, artifact_rows)

    sidecar = PUBLIC / "POST30K_PUBLIC_SAFE_ASSERTION_SIDECAR.tsv"
    route_path = PUBLIC / "POST30K_SOURCE_ROUTE_DISCOVERY.tsv"
    artifact_path = PUBLIC / "POST30K_PUBLIC_ARTIFACT_RECEIPT.tsv"
    identity_path = PUBLIC / "POST30K_COE_IDENTITY_AUDIT.tsv"
    write_tsv(sidecar, safe_rows)
    write_tsv(route_path, route_rows)
    write_tsv(artifact_path, [{
        **row,
        "restricted_relative_path": (
            row["restricted_relative_path"]
            if str(row["restricted_relative_path"]).startswith("restricted-")
            else "restricted://post30k_extension/" + row["restricted_relative_path"]
        ),
    } for row in artifact_rows])
    identity_rows = []
    by_effective: dict[str, list[Any]] = {}
    for atom in coe_atoms:
        by_effective.setdefault(atom.effective_record_id, []).append(atom)
    for effective_id, record_atoms in sorted(by_effective.items()):
        identity_rows.append({
            "post30k_effective_record_id": effective_id,
            "post30k_coffee_identity_id": record_atoms[0].coffee_identity_id,
            "post30k_source_locator": record_atoms[0].source_url,
            "post30k_year_id": f"year.{record_atoms[0].edition_year}",
            "post30k_descriptor_bundle_sha256": hashlib.sha256(
                "\x1f".join(
                    sorted(hashlib.sha256(atom.atomic_text.encode()).hexdigest() for atom in record_atoms)
                ).encode()
            ).hexdigest(),
            "old_domain_candidate_count": 0,
            "candidate_effective_record_id": "",
            "candidate_year_id": "",
            "candidate_name_sha256": "",
            "candidate_descriptor_bundle_sha256": "",
            "name_token_overlap": "NA_NO_CANDIDATE",
            "year_exact": "false",
            "descriptor_bundle_exact": "false",
            "match_state": "NO_CANDIDATE",
            "identity_merge_authorized": "false",
            "decision_basis": "NO_CROSS_DOMAIN_CANDIDATE_FOUND_URL_YEAR_NAME_OR_DESCRIPTOR_BUNDLE_NOT_INFERRED",
        })
    if not identity_rows:
        identity_rows.append({
            "post30k_effective_record_id": "NA_NO_NEW_COE_RECORD",
            "post30k_coffee_identity_id": "",
            "post30k_source_locator": "",
            "post30k_year_id": "",
            "post30k_descriptor_bundle_sha256": "",
            "old_domain_candidate_count": 0,
            "candidate_effective_record_id": "",
            "candidate_year_id": "",
            "candidate_name_sha256": "",
            "candidate_descriptor_bundle_sha256": "",
            "name_token_overlap": "NA_NO_NEW_COE_RECORD",
            "year_exact": "false",
            "descriptor_bundle_exact": "false",
            "match_state": "NO_CANDIDATE",
            "identity_merge_authorized": "false",
            "decision_basis": "NA_NO_NEW_COE_RECORD",
        })
    write_tsv(identity_path, identity_rows)

    non_coe_effort = sum(row["discovery_lane"] == "NON_COE" for row in route_rows)
    total_effort = len(route_rows)
    manifest = {
        "contract_version": "post30k-extension-manifest.v1",
        "extension_batch_id": EXTENSION_BATCH_ID,
        "generated_at": GENERATED_AT,
        "run": True,
        "status": "CLEANED_30K_AND_40K_ACQUISITION_CHECKPOINT_REACHED" if FROZEN_COUNT + accepted >= TARGET_TOTAL else "CLEANED_30K_POST30K_ACQUISITION_PARTIAL",
        "frozen_snapshot_version": "professional-descriptor-candidate-v1-30k",
        "frozen_denominator_assertion_count": FROZEN_COUNT,
        "extension_isolated_from_frozen_snapshot": True,
        "exact_cursor_validated": exact_cursor_validated,
        "net_new_raw_assertion_count": raw_count,
        "net_new_deinflated_assertion_count": accepted,
        "net_new_record_unique_count": record_unique,
        "net_new_effective_record_count": len(effective_records),
        "total_candidate_count": FROZEN_COUNT + accepted,
        "checkpoint_40000_reached": FROZEN_COUNT + accepted >= TARGET_TOTAL,
        "new_coe_assertion_count": sum(atom.counts_as_assertion for atom in coe_atoms),
        "new_non_coe_assertion_count": sum(atom.counts_as_assertion for atom in non_coe_atoms),
        "new_non_coe_positive_family_count": len(positive_non_coe),
        "new_non_coe_positive_families": sorted(positive_non_coe),
        "new_rights_clearable_family_count": len(positive_non_coe),
        "new_non_coe_assertion_target": 1500,
        "new_non_coe_assertion_target_reached": sum(atom.counts_as_assertion for atom in non_coe_atoms) >= 1500,
        "new_non_coe_positive_family_target": 1,
        "new_non_coe_positive_family_target_reached": len(positive_non_coe) >= 1,
        "non_coe_discovery_effort_count": non_coe_effort,
        "total_discovery_effort_count": total_effort,
        "non_coe_discovery_effort_rate": round(non_coe_effort / total_effort, 6),
        "non_coe_discovery_effort_unit": "SOURCE_ROUTE_ATTEMPT",
        "cursor_start": CURSOR_START,
        "cursor_end": cursor_end,
        "coe_route_exhausted": route_exhausted,
        "coe_continuation_blocked": blocked,
        "hard_stop_rule": "FIRST_COMPLETE_EFFECTIVE_RECORD_BOUNDARY_AT_OR_ABOVE_TOTAL_40000",
        "restricted_assertion_ledger_sha256": sha256_file(restricted_ledger),
        "restricted_artifact_receipt_sha256": sha256_file(restricted_artifacts),
        "public_safe_sidecar_sha256": sha256_file(sidecar),
        "model_eligible_assertion_count": 0,
        "human_reviewed_assertion_count": 0,
        "schema_changed": False,
        "new_migration_count": 0,
        "offline_reproduction_policy": "PUBLIC_OUTPUTS_BYTE_IDENTICAL_FROM_RESTRICTED_CACHE",
        "files": [],
    }
    for path in sorted(PUBLIC.glob("*.tsv")):
        manifest["files"].append({
            "path": path.name,
            "sha256": sha256_file(path),
            "byte_count": path.stat().st_size,
            "data_row_count": max(sum(1 for _ in path.open(encoding="utf-8")) - 1, 0),
        })
    manifest_path = PUBLIC / "POST30K_EXTENSION_MANIFEST.json"
    manifest_path.write_text(json.dumps(manifest, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    paths = sorted(path for path in PUBLIC.iterdir() if path.is_file() and path.name != "SHA256SUMS")
    (PUBLIC / "SHA256SUMS").write_text(
        "".join(f"{sha256_file(path)}  {path.name}\n" for path in paths),
        encoding="utf-8",
    )
    print(f"POST30K_EXTENSION_DEINFLATED={accepted}")
    print(f"POST30K_TOTAL_CANDIDATE={FROZEN_COUNT + accepted}")
    print(f"POST30K_NON_COE_ASSERTIONS={manifest['new_non_coe_assertion_count']}")
    print(f"POST30K_NON_COE_POSITIVE_FAMILIES={len(positive_non_coe)}")
    print(f"POST30K_CURSOR_END={cursor_end}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

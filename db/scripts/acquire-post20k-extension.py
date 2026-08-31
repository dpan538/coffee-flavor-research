#!/usr/bin/env python3
"""Continue acquisition from the frozen Batch 2 cursor into isolated staging.

Raw artifacts and source-native text are written only below ``--restricted-root``.
The repository receives a hash-only extension sidecar and route/checkpoint
receipts.  The first complete effective-record boundary at or above 30,000
total candidate assertions is the hard stop.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import importlib.util
import json
import re
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from lxml import html as lxml_html
from openpyxl import load_workbook


# LEGACY_REPRODUCIBILITY_ENTRYPOINT: retained for the accepted 30K snapshot.
LEGACY_REPRODUCIBILITY_ENTRYPOINT = True


ROOT = Path(__file__).resolve().parents[2]
PUBLIC = ROOT / "db" / "data" / "post20k-extension-staging"
FROZEN_RESTRICTED = Path(
    "/private/tmp/round3l-acquisition/professional_descriptor_batch2"
)
FROZEN_COUNT = 20_003
TARGET_TOTAL = 30_000
EXTENSION_TARGET = TARGET_TOTAL - FROZEN_COUNT
START_PAGE = 12
START_INDEX = 10
START_URL = "https://farmdirectory.cupofexcellence.org/listing/9-liquidambar-honduras-2026-parainema-catracha/"
CURSOR_START = f"archive-page={START_PAGE};detail-index={START_INDEX};url={START_URL}"
EXTENSION_BATCH_ID = "professional-descriptor-post20k-extension-20260829"
PARSER_VERSION = "post20k.professional-descriptor-parser.v1"
ADAPTER_VERSION = "post20k.public-safe-adapter.v1"
GENERATED_AT = "2026-08-29T00:00:00Z"

FIGSHARE_API_URL = "https://api.figshare.com/v2/articles/25735122"
FIGSHARE_FILE_URL = "https://ndownloader.figshare.com/files/46039437"
LENGUPA_URL = (
    "https://www.frontiersin.org/journals/sustainable-food-systems/articles/"
    "10.3389/fsufs.2026.1809471/full"
)

LENGUPA_CODES = {
    "1": "citrus", "2": "red fruits", "3": "herbal", "4": "chocolate",
    "5": "nutty", "6": "sweet", "7": "almond", "8": "caramel",
    "9": "panela", "10": "pepper", "11": "honey", "12": "sugarcane",
    "13": "lemongrass", "14": "pulp", "15": "ferment", "16": "earthy",
    "17": "cardboard", "18": "medicinal", "19": "hazelnut", "20": "woody",
    "21": "resting", "22": "banana", "23": "peanut", "24": "fruity",
    "25": "potato", "26": "green bean", "27": "vegetal", "28": "pea",
    "29": "arazá fruit", "30": "straw", "31": "husk", "32": "unresolved code 32",
}


def load_batch2_module():
    path = ROOT / "db" / "scripts" / "acquire-professional-descriptors-batch2.py"
    spec = importlib.util.spec_from_file_location("batch2_acquisition", path)
    if spec is None or spec.loader is None:
        raise RuntimeError("cannot import Batch 2 acquisition primitives")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


B2 = load_batch2_module()


def sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def write_tsv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        raise RuntimeError(f"refusing to write empty TSV: {path}")
    fields = list(rows[0])
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle, fieldnames=fields, delimiter="\t", lineterminator="\n",
            extrasaction="raise",
        )
        writer.writeheader()
        for row in rows:
            writer.writerow({
                key: str(value).lower() if isinstance(value, bool) else value
                for key, value in row.items()
            })


def fetch_retry(url: str, path: Path, *, offline: bool, delay: float) -> bool:
    if path.is_file() and path.stat().st_size:
        return False
    last: Exception | None = None
    for attempt in range(1 if offline else 3):
        try:
            return B2.fetch(url, path, offline=offline, delay=delay if attempt == 0 else 1.0)
        except RuntimeError as exc:
            last = exc
    assert last is not None
    raise last


def source_definition(
    *, family: str, route: str, schema: str, publisher: str,
    rights: str, rights_basis: str, evidence_tier: str, collection_tier: str,
) -> dict[str, str]:
    return {
        "family": family,
        "route": route,
        "schema": schema,
        "publisher": publisher,
        "language": "en",
        "rights": rights,
        "rights_basis": rights_basis,
        "evidence_tier": evidence_tier,
        "collection_tier": collection_tier,
    }


ROBUSTA_SOURCE = source_definition(
    family="family.frontiers_inera_robusta_q_grader_panel",
    route="route.frontiers.2024.inera-robusta-q-grader-frequency",
    schema="schema.research-q-grader.sample-descriptor-frequency.v1",
    publisher="Frontiers / INERA / Coffeelab Independent",
    rights="AFFIRMATIVE",
    rights_basis="CC_BY_4_0_FIGSHARE_DATASET",
    evidence_tier="P2",
    collection_tier="GOLD",
)

LENGUPA_SOURCE = source_definition(
    family="family.frontiers_cenicafe_lengupa_trained_cuppers",
    route="route.frontiers.2026.lengupa-trained-cupper-table",
    schema="schema.research-trained-cupper.sample-coded-descriptors.v1",
    publisher="Frontiers / Cenicafé / CREPIB",
    rights="AFFIRMATIVE",
    rights_basis="CC_BY_FRONTIERS_ARTICLE",
    evidence_tier="P2",
    collection_tier="GOLD",
)


def parse_robusta(path: Path) -> list[list[Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Sensory_scores"]
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(value or "") for value in rows[0]]
    descriptor_columns = range(20, 29)
    artifact_hash = sha256_file(path)
    records: list[list[Any]] = []
    for row_index, values in enumerate(rows[1:], 2):
        genotype = str(values[0] or "").strip()
        year = str(values[1] or "").strip()
        if not genotype or not year:
            continue
        record_atoms: list[Any] = []
        effective = B2.stable_id("effective-post20k", ROBUSTA_SOURCE["route"], genotype, year)
        coffee = B2.stable_id("coffee-post20k", ROBUSTA_SOURCE["family"], genotype)
        for column_index in descriptor_columns:
            frequency = values[column_index]
            if not isinstance(frequency, (int, float)) or frequency <= 0:
                continue
            descriptor = header[column_index]
            record_atoms.extend(B2.make_atoms(
                source=ROBUSTA_SOURCE,
                artifact_sha256=artifact_hash,
                source_url=FIGSHARE_FILE_URL,
                source_locator=f"sheet:Sensory_scores#row={row_index};column={column_index + 1};q-grader-frequency={int(frequency)}",
                effective_record_id=effective,
                coffee_identity_id=coffee,
                edition_or_release="Frontiers INERA Robusta sensory profiles v1",
                edition_year=year,
                preparation_service="FINE_ROBUSTA_STANDARD_CUPPING",
                roast_evidence="STANDARDIZED_SAMPLE_ROAST_PER_ARTICLE",
                source_field_label=descriptor,
                raw_field_text=descriptor,
                publication_layer="TRAINED_PANEL_SAMPLE_CONSENSUS",
                provenance_state="SOURCE_AUDITED_Q_GRADER_FREQUENCY_AGGREGATE",
            ))
        records.append(record_atoms)
    return records


def cell_text(cell: Any) -> str:
    return " ".join(" ".join(cell.itertext()).split())


def parse_lengupa(path: Path) -> list[list[Any]]:
    document = lxml_html.fromstring(path.read_bytes())
    artifact_hash = sha256_file(path)
    table_rows: list[list[str]] = []
    for table in document.xpath("//table"):
        rows = [[cell_text(cell) for cell in tr.xpath("./th|./td")] for tr in table.xpath(".//tr")]
        if rows and any("D.FA" in cell for cell in rows[0]) and any(cell == "D.F" for cell in rows[0]):
            table_rows = rows
            break
    if not table_rows:
        raise RuntimeError("Lengupá sensory Table 3 was not found")
    header = table_rows[0]
    id_index = header.index("ID")
    aroma_index = header.index("D.FA")
    flavor_index = header.index("D.F")
    records: list[list[Any]] = []
    for row_index, values in enumerate(table_rows[1:], 1):
        if len(values) <= max(id_index, aroma_index, flavor_index):
            continue
        sample_id = values[id_index].strip()
        if not re.fullmatch(r"\d+", sample_id):
            continue
        effective = B2.stable_id("effective-post20k", LENGUPA_SOURCE["route"], sample_id)
        coffee = B2.stable_id("coffee-post20k", LENGUPA_SOURCE["family"], sample_id)
        record_atoms: list[Any] = []
        for field_label, column_index in (("Fragrance/aroma descriptors", aroma_index), ("Flavor descriptors", flavor_index)):
            codes = re.findall(r"\d+", values[column_index])
            code_counts = Counter(codes)
            for code in sorted(code_counts, key=int):
                descriptor = LENGUPA_CODES.get(code, f"unresolved code {code}")
                record_atoms.extend(B2.make_atoms(
                    source=LENGUPA_SOURCE,
                    artifact_sha256=artifact_hash,
                    source_url=LENGUPA_URL,
                    source_locator=f"html:Table-3#sample={sample_id};field={column_index};code={code};frequency={code_counts[code]}",
                    effective_record_id=effective,
                    coffee_identity_id=coffee,
                    edition_or_release="Frontiers Lengupá trained-cupper sensory Table 3",
                    edition_year="2026",
                    preparation_service="SCA_GUIDELINE_CUPPING",
                    roast_evidence="STANDARDIZED_CENICAFE_LAB_PREPARATION_PER_ARTICLE",
                    source_field_label=field_label,
                    raw_field_text=descriptor,
                    publication_layer="TRAINED_PANEL_SAMPLE_CONSENSUS",
                    provenance_state="SOURCE_AUDITED_TRAINED_CUPPER_CODED_SAMPLE_DESCRIPTOR",
                ))
        records.append(record_atoms)
    return records


def route_receipt(
    source: dict[str, str], records: list[list[Any]], artifacts: int,
    disposition: str, basis: str,
) -> dict[str, Any]:
    atoms = [atom for record in records for atom in record]
    return {
        "source_family_id": source["family"],
        "publisher": source["publisher"],
        "source_route_id": source["route"],
        "route_schema": source["schema"],
        "discovery_lane": "NON_COE",
        "artifacts_inspected": artifacts,
        "descriptor_bearing_records": len([record for record in records if any(atom.counts_as_assertion for atom in record)]),
        "raw_assertion_count": len(atoms),
        "deinflated_assertion_count": sum(atom.counts_as_assertion for atom in atoms),
        "rights_state": source["rights"],
        "disposition": disposition,
        "basis": basis,
    }


def static_discovery_receipts() -> list[dict[str, Any]]:
    zero = [
        (
            "family.ufla_trained_tds_panel", "Wiley / Federal University of Lavras",
            "route.wiley.2025.trained-panel-tds", "schema.article-aggregate-tds-curves.v1",
            "POSITIVE_AGGREGATE_ONLY_NOT_ROW_LEVEL_STAGED",
            "TRAINED_PANEL_AND_SAMPLE_PROFILES_FOUND_BUT_NO_ROW_LEVEL_OPEN_MATRIX",
        ),
        (
            "family.osu_flavor_research_education_center", "Ohio State / PLOS ONE",
            "route.plos.2019.retronasal-trained-panel", "schema.method-attribute-list.v1",
            "ZERO_YIELD_METHOD_ATTRIBUTE_LIST_ONLY",
            "ATTRIBUTE_LEXICON_AND_METHOD_PRESENT_NO_SAMPLE_OBSERVATION_ROWS",
        ),
        (
            "family.uc_davis_consumer_black_coffee", "UC Davis / Dryad",
            "route.dryad.2023.consumer-black-coffee-cata", "schema.consumer-cata.v1",
            "EXCLUDED_CONSUMER_NOT_PROFESSIONAL",
            "CONSUMER_CATA_OBSERVATIONS_OUTSIDE_PROFESSIONAL_DESCRIPTOR_SCOPE",
        ),
        (
            "family.mendeley_ftir_specialty_coffee", "Mendeley Data",
            "route.mendeley.2025.ftir-sensory-scores", "schema.sensory-score-only.v1",
            "ZERO_YIELD_SCORES_NO_DESCRIPTOR_ROWS",
            "NUMERIC_SENSORY_QUALITY_SCORES_DO_NOT_CREATE_DESCRIPTOR_ASSERTIONS",
        ),
    ]
    return [{
        "source_family_id": family,
        "publisher": publisher,
        "source_route_id": route,
        "route_schema": schema,
        "discovery_lane": "NON_COE",
        "artifacts_inspected": 1,
        "descriptor_bearing_records": 0,
        "raw_assertion_count": 0,
        "deinflated_assertion_count": 0,
        "rights_state": "UNKNOWN",
        "disposition": disposition,
        "basis": basis,
    } for family, publisher, route, schema, disposition, basis in zero]


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--restricted-root", type=Path, required=True)
    parser.add_argument("--offline", action="store_true")
    parser.add_argument("--request-delay", type=float, default=0.2)
    args = parser.parse_args()

    restricted = args.restricted_root / "post20k_extension"
    raw = restricted / "raw"
    raw.mkdir(parents=True, exist_ok=True)
    PUBLIC.mkdir(parents=True, exist_ok=True)
    for path in PUBLIC.iterdir():
        if path.is_file():
            path.unlink()

    frozen_rows = read_tsv(FROZEN_RESTRICTED / "PROFESSIONAL_ASSERTIONS_RESTRICTED.tsv")
    frozen_urls = {row["source_url"] for row in frozen_rows}
    all_atoms: list[Any] = []
    cursor_by_record: dict[str, str] = {}
    artifact_rows: list[dict[str, Any]] = []
    route_rows: list[dict[str, Any]] = []

    # Non-CoE positive route 1: Q-grader frequency aggregates (CC BY 4.0).
    figshare_meta = raw / "non-coe" / "figshare-25735122-metadata.json"
    figshare_xlsx = raw / "non-coe" / "frontiers-inera-robusta-sensory.xlsx"
    fetch_retry(FIGSHARE_API_URL, figshare_meta, offline=args.offline, delay=args.request_delay)
    fetch_retry(FIGSHARE_FILE_URL, figshare_xlsx, offline=args.offline, delay=args.request_delay)
    metadata = json.loads(figshare_meta.read_text(encoding="utf-8"))
    if metadata.get("license", {}).get("name") != "CC BY 4.0":
        raise RuntimeError("Figshare Robusta license drift")
    robusta_records = parse_robusta(figshare_xlsx)
    robusta_atoms = [atom for record in robusta_records for atom in record]
    B2.apply_deinflation(robusta_atoms)
    all_atoms.extend(robusta_atoms)
    for atom in robusta_atoms:
        cursor_by_record[atom.effective_record_id] = "noncoe-route=frontiers-inera-robusta;status=EXHAUSTED"
    route_rows.append(route_receipt(
        ROBUSTA_SOURCE, robusta_records, 2, "POSITIVE_ROUTE_EXHAUSTED",
        "CC_BY_4_0_Q_GRADER_SAMPLE_DESCRIPTOR_FREQUENCIES",
    ))
    for url, path in ((FIGSHARE_API_URL, figshare_meta), (FIGSHARE_FILE_URL, figshare_xlsx)):
        artifact_rows.append({
            "source_route_id": ROBUSTA_SOURCE["route"], "source_url": url,
            "restricted_relative_path": path.relative_to(restricted).as_posix(),
            "sha256": sha256_file(path), "byte_count": path.stat().st_size,
            "acquisition_cursor": "noncoe-route=frontiers-inera-robusta;status=EXHAUSTED",
        })

    # Non-CoE positive route 2: trained Cenicafé cuppers, sample-coded descriptors.
    lengupa_html = raw / "non-coe" / "frontiers-lengupa-trained-cupper-article.html"
    fetch_retry(LENGUPA_URL, lengupa_html, offline=args.offline, delay=args.request_delay)
    lengupa_records = parse_lengupa(lengupa_html)
    lengupa_atoms = [atom for record in lengupa_records for atom in record]
    B2.apply_deinflation(lengupa_atoms)
    all_atoms.extend(lengupa_atoms)
    for atom in lengupa_atoms:
        cursor_by_record[atom.effective_record_id] = "noncoe-route=frontiers-lengupa;status=EXHAUSTED"
    route_rows.append(route_receipt(
        LENGUPA_SOURCE, lengupa_records, 1, "POSITIVE_ROUTE_EXHAUSTED",
        "CC_BY_TRAINED_CENICAFE_CUPPER_SAMPLE_DESCRIPTOR_TABLE",
    ))
    artifact_rows.append({
        "source_route_id": LENGUPA_SOURCE["route"], "source_url": LENGUPA_URL,
        "restricted_relative_path": lengupa_html.relative_to(restricted).as_posix(),
        "sha256": sha256_file(lengupa_html), "byte_count": lengupa_html.stat().st_size,
        "acquisition_cursor": "noncoe-route=frontiers-lengupa;status=EXHAUSTED",
    })
    route_rows.extend(static_discovery_receipts())

    accepted = sum(atom.counts_as_assertion for atom in all_atoms)
    coe_records: list[list[Any]] = []
    cursor_end = CURSOR_START
    blocked = False
    route_exhausted = False
    last_page_urls: list[str] = []
    for page_number in range(START_PAGE, B2.MAX_COE_ARCHIVE_PAGES + 1):
        archive_path = raw / "coe-archive" / f"listings-page-{page_number:03d}.html"
        try:
            fetch_retry(B2.coe_archive_url(page_number), archive_path, offline=args.offline, delay=args.request_delay)
        except RuntimeError:
            cursor_end = f"retry-archive-page={page_number}"
            blocked = True
            break
        urls = sorted(B2.coe_archive_links(archive_path.read_text(encoding="utf-8", errors="replace")))
        artifact_rows.append({
            "source_route_id": "route.coe.archive-continuation",
            "source_url": B2.coe_archive_url(page_number),
            "restricted_relative_path": archive_path.relative_to(restricted).as_posix(),
            "sha256": sha256_file(archive_path), "byte_count": archive_path.stat().st_size,
            "acquisition_cursor": f"archive-page={page_number}",
        })
        if not urls:
            cursor_end = f"archive-page={page_number};NO_LISTING_LINKS"
            blocked = True
            break
        last_page_urls = urls
        first_index = START_INDEX if page_number == START_PAGE else 1
        if page_number == START_PAGE:
            if len(urls) < START_INDEX or urls[START_INDEX - 1] != START_URL:
                raise RuntimeError("preserved continuation cursor no longer matches archive page 12")
        for detail_index, source_url in enumerate(urls, 1):
            if detail_index < first_index:
                continue
            if source_url in frozen_urls:
                # Archive pagination is live and may shift between runs.  The
                # exact page-12/detail-10 boundary is verified above; later
                # frozen URLs are publication overlaps and are skipped rather
                # than reacquired or credited again.
                continue
            if FROZEN_COUNT + accepted >= TARGET_TOTAL:
                cursor_end = f"archive-page={page_number};detail-index={detail_index};url={source_url}"
                break
            detail_path = raw / "coe-detail" / f"{hashlib.sha256(source_url.encode()).hexdigest()[:24]}.html"
            try:
                fetch_retry(source_url, detail_path, offline=args.offline, delay=args.request_delay)
            except RuntimeError:
                cursor_end = f"retry-archive-page={page_number};detail-index={detail_index};url={source_url}"
                blocked = True
                break
            atoms = B2.parse_coe_detail(detail_path, source_url)
            B2.apply_deinflation(atoms)
            if atoms:
                coe_records.append(atoms)
                all_atoms.extend(atoms)
                accepted += sum(atom.counts_as_assertion for atom in atoms)
                cursor = f"archive-page={page_number};detail-index={detail_index};url={source_url}"
                for atom in atoms:
                    cursor_by_record[atom.effective_record_id] = cursor
                artifact_rows.append({
                    "source_route_id": atoms[0].source_route, "source_url": source_url,
                    "restricted_relative_path": detail_path.relative_to(restricted).as_posix(),
                    "sha256": sha256_file(detail_path), "byte_count": detail_path.stat().st_size,
                    "acquisition_cursor": cursor,
                })
            if len(coe_records) % 50 == 0 and coe_records:
                print(
                    f"POST20K_COE_PROGRESS records={len(coe_records)} extension={accepted} total={FROZEN_COUNT + accepted}",
                    flush=True,
                )
        if blocked or FROZEN_COUNT + accepted >= TARGET_TOTAL:
            if FROZEN_COUNT + accepted >= TARGET_TOTAL and not cursor_end.startswith("archive-page="):
                next_index = len(last_page_urls) + 1
                cursor_end = f"archive-page={page_number};detail-index={next_index};END_OF_PAGE"
            break
    else:
        route_exhausted = True
        cursor_end = "EXHAUSTED_COE_ARCHIVE"

    coe_source = {
        **B2.COE,
        "route": "route.coe.post20k-archive-continuation",
        "schema": "schema.coe.structured-official-sensory-fields.v2",
        "evidence_tier": "MIXED",
        "collection_tier": "MIXED",
    }
    coe_atoms = [atom for record in coe_records for atom in record]
    route_rows.append({
        "source_family_id": B2.COE["family"],
        "publisher": B2.COE["publisher"],
        "source_route_id": "route.coe.post20k-archive-continuation",
        "route_schema": "schema.coe.mixed-explicit-jury-and-generic.v2",
        "discovery_lane": "COE_CONTINUATION",
        "artifacts_inspected": len(coe_records),
        "descriptor_bearing_records": len(coe_records),
        "raw_assertion_count": len(coe_atoms),
        "deinflated_assertion_count": sum(atom.counts_as_assertion for atom in coe_atoms),
        "rights_state": "MIXED_PENDING_UNKNOWN",
        "disposition": "POSITIVE_ROUTE_CHECKPOINT_REACHED" if FROZEN_COUNT + accepted >= TARGET_TOTAL else "PARTIAL_OR_EXHAUSTED",
        "basis": "EXACT_BATCH2_CURSOR_CONTINUATION_FIRST_COMPLETE_RECORD_BOUNDARY",
    })

    # Re-run de-inflation across the entire extension to make offline output independent
    # of per-route parsing order while preserving complete record boundaries.
    B2.apply_deinflation(all_atoms)
    accepted = sum(atom.counts_as_assertion for atom in all_atoms)
    raw_count = len(all_atoms)
    record_unique = sum(atom.counts_as_record_unique_descriptor for atom in all_atoms)
    effective_records = {atom.effective_record_id for atom in all_atoms if atom.counts_as_assertion}
    non_coe_atoms = [atom for atom in all_atoms if atom.source_family != B2.COE["family"]]
    coe_atoms = [atom for atom in all_atoms if atom.source_family == B2.COE["family"]]
    positive_non_coe = {
        atom.source_family for atom in non_coe_atoms if atom.counts_as_assertion
    }

    safe_rows: list[dict[str, Any]] = []
    for atom, safe in zip(all_atoms, B2.safe_rows(all_atoms)):
        safe_rows.append({
            "extension_batch_id": EXTENSION_BATCH_ID,
            **safe,
            # Some source schemas use a descriptor as the column heading.  Hash
            # every field label at this bridge instead of assuming it is
            # non-content metadata.
            "source_field_label": (
                "hash:sha256:"
                + hashlib.sha256(atom.source_field_label.encode("utf-8")).hexdigest()
            ),
            "parser_version": PARSER_VERSION,
            "adapter_version": ADAPTER_VERSION,
            "acquisition_cursor": cursor_by_record[atom.effective_record_id],
            "frozen_snapshot_version": "professional-descriptor-candidate-v0-20k",
            "frozen_snapshot_member": False,
        })
    restricted_rows = list(B2.restricted_rows(all_atoms))
    restricted_ledger = restricted / "POST20K_ASSERTIONS_RESTRICTED.tsv"
    restricted_artifacts = restricted / "POST20K_RAW_ARTIFACT_RECEIPT.tsv"
    B2.write_tsv(restricted_ledger, restricted_rows)
    write_tsv(restricted_artifacts, artifact_rows)

    sidecar = PUBLIC / "POST20K_PUBLIC_SAFE_ASSERTION_SIDECAR.tsv"
    route_path = PUBLIC / "POST20K_SOURCE_ROUTE_DISCOVERY.tsv"
    artifact_path = PUBLIC / "POST20K_PUBLIC_ARTIFACT_RECEIPT.tsv"
    write_tsv(sidecar, safe_rows)
    write_tsv(route_path, route_rows)
    public_artifacts = [{
        **row,
        "restricted_relative_path": "restricted://post20k_extension/" + row["restricted_relative_path"],
    } for row in artifact_rows]
    write_tsv(artifact_path, public_artifacts)

    non_coe_effort_count = sum(row["discovery_lane"] == "NON_COE" for row in route_rows)
    discovery_effort_count = len(route_rows)
    manifest = {
        "contract_version": "post20k-extension-manifest.v1",
        "extension_batch_id": EXTENSION_BATCH_ID,
        "generated_at": GENERATED_AT,
        "run": True,
        "status": (
            "CLEANING_PASS_30K_CANDIDATE_CHECKPOINT_REACHED"
            if FROZEN_COUNT + accepted >= TARGET_TOTAL
            else "CLEANING_PASS_COE_ROUTE_EXHAUSTED"
            if route_exhausted
            else "CLEANING_PASS_NON_COE_DIVERSIFICATION_GAPS"
        ),
        "frozen_snapshot_version": "professional-descriptor-candidate-v0-20k",
        "frozen_denominator_assertion_count": FROZEN_COUNT,
        "extension_isolated_from_frozen_snapshot": True,
        "net_new_raw_assertion_count": raw_count,
        "net_new_deinflated_assertion_count": accepted,
        "net_new_record_unique_count": record_unique,
        "net_new_effective_record_count": len(effective_records),
        "total_candidate_count": FROZEN_COUNT + accepted,
        "checkpoint_30000_reached": FROZEN_COUNT + accepted >= TARGET_TOTAL,
        "new_coe_assertion_count": sum(atom.counts_as_assertion for atom in coe_atoms),
        "new_non_coe_assertion_count": sum(atom.counts_as_assertion for atom in non_coe_atoms),
        "new_non_coe_positive_family_count": len(positive_non_coe),
        "new_non_coe_positive_families": sorted(positive_non_coe),
        "new_non_coe_positive_family_target": 3,
        "new_non_coe_assertion_target": 3000,
        "non_coe_discovery_effort_count": non_coe_effort_count,
        "total_discovery_effort_count": discovery_effort_count,
        "non_coe_discovery_effort_rate": round(non_coe_effort_count / discovery_effort_count, 6),
        "non_coe_discovery_effort_unit": "SOURCE_ROUTE_ATTEMPT",
        "cursor_start": CURSOR_START,
        "cursor_end": cursor_end,
        "coe_route_exhausted": route_exhausted,
        "coe_continuation_blocked": blocked,
        "hard_stop_rule": "FIRST_COMPLETE_EFFECTIVE_RECORD_BOUNDARY_AT_OR_ABOVE_TOTAL_30000",
        "restricted_assertion_ledger_sha256": sha256_file(restricted_ledger),
        "restricted_artifact_receipt_sha256": sha256_file(restricted_artifacts),
        "public_safe_sidecar_sha256": sha256_file(sidecar),
        "model_eligible_assertion_count": 0,
        "human_reviewed_assertion_count": 0,
        "schema_changed": False,
        "new_migration_count": 0,
        "offline_reproduction_policy": "PUBLIC_OUTPUTS_BYTE_IDENTICAL_FROM_RESTRICTED_CACHE",
        "files": [],
    }
    for path in sorted(PUBLIC.glob("*.tsv")):
        manifest["files"].append({
            "path": path.name, "sha256": sha256_file(path),
            "byte_count": path.stat().st_size,
            "data_row_count": max(sum(1 for _ in path.open(encoding="utf-8")) - 1, 0),
        })
    manifest_path = PUBLIC / "POST20K_EXTENSION_MANIFEST.json"
    manifest_path.write_text(json.dumps(manifest, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    paths = sorted(path for path in PUBLIC.iterdir() if path.is_file() and path.name != "SHA256SUMS")
    (PUBLIC / "SHA256SUMS").write_text(
        "".join(f"{sha256_file(path)}  {path.name}\n" for path in paths),
        encoding="utf-8",
    )
    print(f"POST20K_EXTENSION_DEINFLATED={accepted}")
    print(f"POST20K_TOTAL_CANDIDATE={FROZEN_COUNT + accepted}")
    print(f"POST20K_NON_COE_ASSERTIONS={manifest['new_non_coe_assertion_count']}")
    print(f"POST20K_NON_COE_POSITIVE_FAMILIES={len(positive_non_coe)}")
    print(f"POST20K_CURSOR_END={cursor_end}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
